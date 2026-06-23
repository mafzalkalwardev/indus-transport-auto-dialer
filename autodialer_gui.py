#!/usr/bin/env python3
"""
FT Solutions — Auto Dialer Pro
Google Voice runs silently inside an embedded browser.
Agents see the FT Solutions interface and the live Google Voice panel only when needed.
"""
import os
import sys
import json
import argparse
import traceback
from datetime import datetime, timedelta

from src.webengine_env import configure_webengine_environment

configure_webengine_environment()

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QDialog, QVBoxLayout,
    QHBoxLayout, QLabel, QPushButton, QLineEdit, QStackedWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QSpinBox, QDoubleSpinBox, QFileDialog, QMessageBox,
    QTextEdit, QFrame, QProgressBar, QScrollArea, QSizePolicy,
    QTabWidget, QSplitter, QGroupBox, QRadioButton, QButtonGroup,
    QFormLayout, QAbstractItemView, QMenu, QGridLayout, QCheckBox,
)
from PyQt6.QtCore import (
    Qt, QTimer, QSize, QUrl, pyqtSignal, QThread, QObject, QLockFile,
    QPoint,
)
from PyQt6.QtGui import (
    QColor, QPalette, QFont, QIcon, QPixmap, QAction, QFontDatabase,
)
from PyQt6.QtWebEngineWidgets import QWebEngineView

import pandas as pd

from src.paths       import (LOGO_PNG, LOGO_JPEG, CONFIG_FILE,
                              CHROME_PROFILES_DIR, LOGS_DIR)
from src.crm_db      import CRMDatabase
from src.phone_utils import clean_phone, fmt_e164, fmt_display
from src.gv_controller import (
    GVController,
    has_session_marker,
    write_session_marker,
)
from src.ui_theme import (
    DARK_QSS, LIGHT_QSS, DEFAULT_THEME,
    status_label, status_color,
    apply_theme, table_item_color, log_status_color,
)
from src.client_deploy import export_client_package, is_client_deployment
from src.slot_watchdog import SlotWatchdog, webengine_total_memory_mb
from src.retry_queue import DialRetryQueue
from src.dialer_logging import setup_dialer_logging, log_info, log_warning, log_error, log_path
from src.process_cleanup import cleanup_stale_webengine_processes
from src.system_profile import (
    recommended_slots, low_resource_reason, system_ram_gb, chrome_process_count,
    effective_enable_ai_audio,
)
from src.pacing.engine import PredictivePacingEngine, PacingMetrics, PacingConfig
from src.pacing.agents import available_agent_count
from src.gv_accounts import (
    load_accounts as load_gv_accounts,
    save_accounts as save_gv_accounts,
    make_profile_name,
    profile_dir as gv_profile_dir,
    clone_profile_folder,
    has_session_marker as gv_has_session_marker,
)

try:
    from PIL import Image
    PIL_OK = True
except ImportError:
    PIL_OK = False

# ── Constants ─────────────────────────────────────────────────────────────────
APP_NAME     = "FT Solutions — Auto Dialer Pro"
WHATSAPP_URL = "https://wa.me/923079670503"
WA_NUMBER    = "+92 307 967 0503"

UI_CONNECTED_STATES = {
    "CONNECTED",
    "ANSWERED",
    "CONNECTED_AUDIO_EVIDENCE",
    "CONNECTED_MANUAL_CONFIRMATION",
    "HUMAN",
}
UI_STATE_DISPLAY = {
    "IDLE": "IDLE",
    "READY": "READY",
    "SETUP REQUIRED": "SETUP REQUIRED",
    "LOADING": "LOADING",
    "LOGIN_NEEDED": "LOGIN_NEEDED",
    "DIALING": "DIALING",
    "CALLING": "DIALING",
    "RINGING": "RINGING",
    "ANSWERED_PENDING": "ANSWERED_PENDING",
    "CONNECTED_CTRL": "ANSWERED_PENDING",
    "NO_ANSWER": "NO_ANSWER",
    "FAILED": "FAILED",
    "ENDED": "ENDED",
    "ENDED_MANUALLY": "ENDED",
    "BUSY": "BUSY",
    "VOICEMAIL": "VOICEMAIL",
}
UI_STATE_ORDER = {
    "IDLE": 0,
    "READY": 0,
    "SETUP REQUIRED": 0,
    "LOADING": 1,
    "LOGIN_NEEDED": 1,
    "DIALING": 2,
    "RINGING": 3,
    "ANSWERED_PENDING": 4,
    "CONNECTED": 5,
    "VOICEMAIL": 5,
    "BUSY": 5,
    "NO_ANSWER": 6,
    "FAILED": 6,
    "ENDED": 7,
}


def ui_display_state(state: str) -> str:
    raw = str(state or "IDLE").upper()
    if raw in UI_CONNECTED_STATES:
        return "CONNECTED"
    return UI_STATE_DISPLAY.get(raw, raw)


def ui_state_allows_transition(current: str, incoming: str) -> bool:
    current_display = ui_display_state(current)
    incoming_display = ui_display_state(incoming)
    if current_display == "CONNECTED" and incoming_display in {
        "DIALING",
        "RINGING",
        "UNKNOWN",
        "FAILED",
        "VOICEMAIL",
        "NO_ANSWER",
    }:
        return False
    if current_display == "CONNECTED" and incoming_display in {"ENDED", "IDLE"}:
        return True
    return (
        UI_STATE_ORDER.get(incoming_display, 0) >= UI_STATE_ORDER.get(current_display, 0)
        or incoming_display in {"IDLE", "READY", "SETUP REQUIRED"}
    )


def _load_dotenv_file(path: str = ".env") -> None:
    if not os.path.exists(path):
        return
    try:
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and key not in os.environ:
                    os.environ[key] = value
    except Exception:
        log_error("Failed to load .env\n" + traceback.format_exc())


def _admin_env_credentials() -> tuple[str, str]:
    return (
        os.environ.get("ADMIN_EMAIL", "").strip(),
        os.environ.get("ADMIN_PASSWORD", ""),
    )

# ── Config ────────────────────────────────────────────────────────────────────
def _load_cfg() -> dict:
    defaults = {
        "theme": DEFAULT_THEME,
        "n_slots": 1,
        "call_timeout": 60,
        "cooldown": 4.0,
        "dial_stagger_sec": 0.8,
        "voicemail_hangup_sec": 3,
        "excel_path": "",
        "deployment_mode": "admin",
        "max_retries": 3,
        "retry_backoff_sec": [5, 15, 45],
        "watchdog_heartbeat_timeout_sec": 45,
        "watchdog_stuck_state_sec": 90,
        "slot_memory_limit_mb": 1200,
        "slot_recycle_after_calls": 75,
        "watchdog_check_interval_sec": 5,
        "enable_ai_audio": True,
        "amd_mode": "heuristic",
        "amd_early_decision_ms": 800,
        "amd_max_decision_ms": 2500,
        "amd_human_first_seconds": 5,
        "audio_device": "",
        "live_debug_mode": False,
        "dry_run_mode": False,
        "predictive_mode": False,
        "max_concurrent_dials": 3,
        "target_agent_wait_seconds": 3,
        "pacing_interval_sec": 2.5,
        "websocket_enabled": False,
        "websocket_port": 8765,
        "ui": {"auto_open_panel_on_human": False},
    }
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f:
                d = json.load(f)
            for k, v in defaults.items():
                if k not in d:
                    d[k] = v
            return d
        except Exception:
            pass
    return defaults

def _save_cfg(cfg: dict) -> None:
    try:
        with open(CONFIG_FILE, "w") as f:
            json.dump(cfg, f, indent=2)
    except Exception:
        pass

# ── Helpers ───────────────────────────────────────────────────────────────────
def _icon() -> QIcon:
    for path in (LOGO_PNG, LOGO_JPEG):
        if os.path.exists(path):
            return QIcon(path)
    return QIcon()

def _pixmap(h: int = 48) -> QPixmap | None:
    for path in (LOGO_PNG, LOGO_JPEG):
        if os.path.exists(path):
            px = QPixmap(path)
            if not px.isNull():
                return px.scaledToHeight(h, Qt.TransformationMode.SmoothTransformation)
    return None

def _btn(text: str, obj_name: str = "", parent=None) -> QPushButton:
    b = QPushButton(text, parent)
    if obj_name:
        b.setObjectName(obj_name)
    return b

def _label(text: str, obj_name: str = "", bold: bool = False,
           size: int = 0, parent=None) -> QLabel:
    lbl = QLabel(text, parent)
    if obj_name:
        lbl.setObjectName(obj_name)
    if bold or size:
        f = lbl.font()
        if bold:
            f.setBold(True)
        if size:
            f.setPointSize(size)
        lbl.setFont(f)
    return lbl

def _hline() -> QFrame:
    line = QFrame()
    line.setObjectName("hline")
    line.setFrameShape(QFrame.Shape.HLine)
    return line


def _configure_data_table(table: QTableWidget, min_visible_rows: int = 5) -> None:
    """Readable row height and minimum size for data tables."""
    table.setMinimumHeight(min_visible_rows * 38 + 48)
    vh = table.verticalHeader()
    vh.setDefaultSectionSize(36)
    vh.setMinimumSectionSize(32)
    table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
    table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)


def _table_item(text: str, theme: str = DEFAULT_THEME) -> QTableWidgetItem:
    item = QTableWidgetItem(str(text))
    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
    item.setForeground(QColor(table_item_color(theme)))
    return item


# ══════════════════════════════════════════════════════════════════════════════
#  SLOT CARD WIDGET
# ══════════════════════════════════════════════════════════════════════════════

class SlotCard(QGroupBox):
    next_clicked = pyqtSignal(int)
    cut_clicked = pyqtSignal(int)
    listen_clicked = pyqtSignal(int)
    monitor_clicked = pyqtSignal(int)
    test_clicked = pyqtSignal(int)
    redial_clicked = pyqtSignal(int)

    MIN_WIDTH = 248

    def __init__(self, slot_id: int, parent=None):
        super().__init__("", parent)
        self.setObjectName("slotCard")
        self.slot_id = slot_id
        self.setMinimumWidth(self.MIN_WIDTH)
        self.setSizePolicy(
            QSizePolicy.Policy.MinimumExpanding,
            QSizePolicy.Policy.Preferred,
        )
        self._build()

    def set_line_label(self, text: str) -> None:
        self.lbl_title.setText(text.strip() or f"Line {self.slot_id + 1}")

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setSpacing(8)
        lay.setContentsMargins(4, 4, 4, 4)

        self._current_state = "IDLE"
        self._gv_ready = False

        self.lbl_title = _label(f"Line {self.slot_id + 1}", bold=True, size=11)
        self.lbl_title.setWordWrap(True)
        lay.addWidget(self.lbl_title)

        self.lbl_status = _label("Setup required", bold=True)
        self._apply_status_style("SETUP REQUIRED")
        lay.addWidget(self.lbl_status)

        self.lbl_phone = _label("No active number", "muted")
        self.lbl_phone.setWordWrap(True)
        lay.addWidget(self.lbl_phone)

        self.lbl_dial_detail = _label("", "muted")
        self.lbl_dial_detail.setWordWrap(True)
        lay.addWidget(self.lbl_dial_detail)

        self.lbl_pickup = _label("", "accent", bold=True)
        self.lbl_pickup.setWordWrap(True)
        self.lbl_pickup.setMinimumHeight(28)
        lay.addWidget(self.lbl_pickup)

        self.lbl_dur = _label("Call time: —", "muted")
        lay.addWidget(self.lbl_dur)

        self.lbl_ai_audio = _label("AI Audio: OFF", "muted")
        self.lbl_ai_audio.setWordWrap(True)
        lay.addWidget(self.lbl_ai_audio)

        self.lbl_ai_decision = _label("Decision: --", "muted")
        self.lbl_ai_decision.setWordWrap(True)
        lay.addWidget(self.lbl_ai_decision)

        self.lbl_external_provider = _label("Ext AMD: off", "muted")
        self.lbl_external_provider.setWordWrap(True)
        lay.addWidget(self.lbl_external_provider)

        btn_col = QVBoxLayout()
        btn_col.setSpacing(6)

        self.btn_next = _btn("Release slot / next", "green")
        self.btn_next.setEnabled(False)
        self.btn_next.setMinimumHeight(36)
        self.btn_next.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.btn_next.clicked.connect(lambda: self.next_clicked.emit(self.slot_id))
        btn_col.addWidget(self.btn_next)

        self.btn_cut = _btn("End call", "red")
        self.btn_cut.setEnabled(False)
        self.btn_cut.setMinimumHeight(36)
        self.btn_cut.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.btn_cut.clicked.connect(lambda: self.cut_clicked.emit(self.slot_id))
        btn_col.addWidget(self.btn_cut)

        self.btn_listen = _btn("Listen", "secondary")
        self.btn_listen.setMinimumHeight(36)
        self.btn_listen.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.btn_listen.setToolTip(
            "Available only after a call is picked up")
        self.btn_listen.setEnabled(False)
        self.btn_listen.clicked.connect(
            lambda: self.listen_clicked.emit(self.slot_id))
        btn_col.addWidget(self.btn_listen)

        self.btn_monitor = _btn("Monitor audio", "secondary")
        self.btn_monitor.setMinimumHeight(36)
        self.btn_monitor.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.btn_monitor.setToolTip(
            "Debug/test: listen to this line even while ringing or voicemail is playing")
        self.btn_monitor.clicked.connect(
            lambda: self.monitor_clicked.emit(self.slot_id))
        btn_col.addWidget(self.btn_monitor)

        self.btn_test = _btn("Test call", "secondary")
        self.btn_test.setMinimumHeight(36)
        self.btn_test.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.btn_test.setToolTip(
            "Place one manual test call on this line without starting a campaign")
        self.btn_test.clicked.connect(lambda: self.test_clicked.emit(self.slot_id))
        btn_col.addWidget(self.btn_test)

        self.btn_redial = _btn("Retry dial", "secondary")
        self.btn_redial.setMinimumHeight(36)
        self.btn_redial.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.btn_redial.setToolTip(
            "Number entered but call did not start — click Call again in Google Voice")
        self.btn_redial.setEnabled(False)
        self.btn_redial.clicked.connect(
            lambda: self.redial_clicked.emit(self.slot_id))
        btn_col.addWidget(self.btn_redial)
        lay.addLayout(btn_col)

    def set_dial_detail(self, text: str) -> None:
        self.lbl_dial_detail.setText(text.strip())

    def _apply_status_style(self, key: str) -> None:
        c = status_color(key)
        self.lbl_status.setStyleSheet(
            f"color: {c}; font-weight: 600; font-size: 11pt;")

    def update_state(self, state: str, phone: str = "", elapsed: str = ""):
        display_state = ui_display_state(state)
        if not ui_state_allows_transition(getattr(self, "_current_state", "IDLE"), display_state):
            return
        self._current_state = display_state
        if display_state == "IDLE":
            display_state = "READY" if self._gv_ready else "SETUP REQUIRED"
        self.lbl_status.setText(status_label(display_state))
        self._apply_status_style(display_state)
        if display_state == "CONNECTED":
            phone_text = phone if phone else "Answered call"
            self.lbl_pickup.setText("CALL PICKED UP - talk now")
        elif display_state in ("DIALING", "RINGING", "VOICEMAIL", "BUSY"):
            phone_text = "Screening in background"
            self.lbl_pickup.setText("")
        else:
            phone_text = "No active number"
            self.lbl_pickup.setText("")
        self.lbl_phone.setText(phone_text)
        self.lbl_dur.setText(
            f"Call time: {elapsed}" if elapsed else "Call time: —")
        active = display_state in ("DIALING", "RINGING", "CONNECTED", "VOICEMAIL", "BUSY")
        self.btn_next.setEnabled(display_state == "CONNECTED")
        self.btn_cut.setEnabled(active)
        self.btn_listen.setEnabled(display_state == "CONNECTED")
        self.btn_monitor.setEnabled(display_state in ("DIALING", "RINGING", "CONNECTED", "VOICEMAIL", "BUSY"))
        self.btn_test.setEnabled(not active)
        self.btn_redial.setEnabled(display_state == "DIALING")

        self.setProperty("connected", display_state == "CONNECTED")
        self.style().unpolish(self)
        self.style().polish(self)

    def set_gv_login_ready(self, ready: bool) -> None:
        self._gv_ready = ready
        if getattr(self, "_current_state", "IDLE") != "IDLE":
            return
        key = "READY" if ready else "SETUP REQUIRED"
        self.lbl_status.setText(status_label(key))
        self._apply_status_style(key)

    def update_detection(self, debug: dict) -> None:
        backend = str(debug.get("audio_backend") or "OFF")
        if backend == "ON":
            self.lbl_ai_audio.setText("AI Audio: ON")
        elif backend == "NO_BACKEND":
            self.lbl_ai_audio.setText("AI Audio: NO_BACKEND")
        else:
            self.lbl_ai_audio.setText("AI Audio: OFF")
        fused = str(debug.get("fused_state") or debug.get("ui_state") or "--")
        conf = debug.get("confidence", 0)
        reason = str(debug.get("reason") or debug.get("audio_reason") or "")
        latency = int(debug.get("detection_time_ms") or 0)
        if len(reason) > 72:
            reason = reason[:69] + "..."
        latency_txt = f" {latency}ms" if latency > 0 else ""
        self.lbl_ai_decision.setText(f"Decision: {fused} ({conf}){latency_txt} {reason}")
        self.set_dial_detail(f"AMD: {fused}{latency_txt} — {reason}" if reason else f"AMD: {fused}{latency_txt}")

        ext_enabled = bool(debug.get("external_detector_enabled", False))
        ext_mode = str(debug.get("external_detector_mode", "off"))
        ext_health = str(debug.get("external_provider_health", "unknown"))
        ext_label = str(debug.get("external_last_label", "none"))
        ext_conf = float(debug.get("external_confidence", 0.0) or 0.0)
        if ext_enabled:
            health_color = "green" if ext_health == "connected" else "red" if ext_health == "error" else "orange"
            self.lbl_external_provider.setText(
                f"Ext {ext_mode}: {ext_health} | {ext_label} ({ext_conf:.2f})"
            )
            self.lbl_external_provider.setStyleSheet(f"color: {health_color}; font-size: 9pt;")
        else:
            self.lbl_external_provider.setText("Ext AMD: off")
            self.lbl_external_provider.setStyleSheet("color: gray; font-size: 9pt;")

    def set_ai_audio_enabled(self, enabled: bool) -> None:
        self.lbl_ai_audio.setText("AI Audio: ON" if enabled else "AI Audio: OFF")


# ══════════════════════════════════════════════════════════════════════════════
#  ADMIN SETUP PAGE
# ══════════════════════════════════════════════════════════════════════════════

class AdminSetupPage(QWidget):
    done = pyqtSignal()

    def __init__(self, db: CRMDatabase, parent=None):
        super().__init__(parent)
        self.db = db
        lay = QVBoxLayout(self)
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # Logo
        px = _pixmap(64)
        if px:
            lbl_logo = QLabel()
            lbl_logo.setPixmap(px)
            lbl_logo.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lay.addWidget(lbl_logo)

        lay.addWidget(_label("FT Solutions", bold=True, size=16,
                              parent=self))
        lay.addWidget(_label("Create your Administrator account to get started",
                              "muted", parent=self))
        lay.addSpacing(20)

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        self.e_name  = QLineEdit(); self.e_name.setPlaceholderText("Full name")
        self.e_email = QLineEdit(); self.e_email.setPlaceholderText("admin@company.com")
        self.e_pw    = QLineEdit(); self.e_pw.setEchoMode(QLineEdit.EchoMode.Password)
        self.e_pw.setPlaceholderText("Min. 8 characters")
        self.e_pw2   = QLineEdit(); self.e_pw2.setEchoMode(QLineEdit.EchoMode.Password)
        self.e_pw2.setPlaceholderText("Repeat password")
        for w in (self.e_name, self.e_email, self.e_pw, self.e_pw2):
            w.setMinimumWidth(300)
        form.addRow("Full Name:", self.e_name)
        form.addRow("Email:",     self.e_email)
        form.addRow("Password:",  self.e_pw)
        form.addRow("Confirm:",   self.e_pw2)
        lay.addLayout(form)
        lay.addSpacing(16)

        btn = _btn("Create Admin Account", "green")
        btn.setMinimumWidth(240)
        btn.clicked.connect(self._submit)
        lay.addWidget(btn, alignment=Qt.AlignmentFlag.AlignCenter)

    def _submit(self):
        name  = self.e_name.get() if hasattr(self.e_name, 'get') else self.e_name.text().strip()
        email = self.e_email.text().strip()
        pw    = self.e_pw.text()
        pw2   = self.e_pw2.text()
        if not all([name, email, pw]):
            QMessageBox.warning(self, "Missing Fields", "All fields are required.")
            return
        if pw != pw2:
            QMessageBox.warning(self, "Mismatch", "Passwords do not match.")
            return
        if len(pw) < 8:
            QMessageBox.warning(self, "Weak Password", "Use at least 8 characters.")
            return
        try:
            self.db.create_admin(email, name, pw)
            QMessageBox.information(
                self, "Account Created",
                f"Admin account created!\n\nEmail: {email}\n\n"
                "Keep these credentials secure. This setup will not appear again."
            )
            self.done.emit()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


# ══════════════════════════════════════════════════════════════════════════════
#  LOGIN PAGE
# ══════════════════════════════════════════════════════════════════════════════

class ClientNotConfiguredPage(QWidget):
    """Shown on client PCs that were not prepared with an agent account."""

    def __init__(self, parent=None):
        super().__init__(parent)
        lay = QVBoxLayout(self)
        lay.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(_label("Workstation not configured", "heroTitle"))
        msg = _label(
            "This copy of the dialer is set up for agents only, but no user "
            "account was found.\n\n"
            "Your administrator must export a client package from their "
            "computer (Administration → Export client package) and copy "
            "the logs and data folders onto this PC.",
            "muted",
        )
        msg.setWordWrap(True)
        msg.setMaximumWidth(480)
        lay.addWidget(msg)


class LoginPage(QWidget):
    login_success = pyqtSignal(dict)

    def __init__(self, db: CRMDatabase, client_mode: bool = False, parent=None):
        super().__init__(parent)
        self.setObjectName("loginPage")
        self.db = db
        self._client_mode = client_mode
        outer = QVBoxLayout(self)
        outer.setAlignment(Qt.AlignmentFlag.AlignCenter)

        card = QFrame()
        card.setObjectName("loginCard")
        card.setFixedWidth(420)
        lay = QVBoxLayout(card)
        lay.setSpacing(12)
        lay.setContentsMargins(36, 32, 36, 32)

        px = _pixmap(64)
        if px:
            lbl = QLabel()
            lbl.setPixmap(px)
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lay.addWidget(lbl)
            lay.addSpacing(4)

        lay.addWidget(_label("FT Solutions", "brandName", bold=True, size=16))
        if client_mode:
            lay.addWidget(_label("Agent sign-in", "accent"))
        else:
            lay.addWidget(_label("Sign in to your dialer account", "muted"))
        lay.addSpacing(16)

        self.e_email = QLineEdit()
        self.e_email.setPlaceholderText("Work email")
        lay.addWidget(self.e_email)
        self.e_pw = QLineEdit()
        self.e_pw.setEchoMode(QLineEdit.EchoMode.Password)
        self.e_pw.setPlaceholderText("Password")
        self.e_pw.returnPressed.connect(self._login)
        lay.addWidget(self.e_pw)
        lay.addSpacing(8)

        self.lbl_err = _label("", "danger")
        lay.addWidget(self.lbl_err, alignment=Qt.AlignmentFlag.AlignCenter)

        btn = _btn("Sign in", "primary")
        btn.setMinimumHeight(42)
        btn.clicked.connect(self._login)
        lay.addWidget(btn)
        lay.addSpacing(8)

        lay.addWidget(
            _label("Need access? Contact your administrator.", "muted"),
            alignment=Qt.AlignmentFlag.AlignCenter,
        )
        outer.addWidget(card)

    def _login(self):
        email = self.e_email.text().strip()
        user = self.db.authenticate(email, self.e_pw.text())
        if not user:
            log_warning(f"Login failed for {email or '<blank>'}")
            self.lbl_err.setText("Incorrect email or password.")
            return
        if self._client_mode and user.get("role") == "admin":
            log_warning(f"Admin login blocked on client workstation for {email}")
            self.lbl_err.setText(
                "This PC is for agents only. Use the agent login your "
                "administrator gave you.")
            return
        log_info(f"Login success for {user.get('email', email)}")
        self.login_success.emit(user)


# ══════════════════════════════════════════════════════════════════════════════
#  GV SETUP DIALOG  (shown when profile not yet logged in)
# ══════════════════════════════════════════════════════════════════════════════

class GVSetupDialog(QDialog):
    """Shows the embedded browser so user can log into Google Voice."""

    login_succeeded = pyqtSignal()

    def __init__(self, controller: GVController, account_label: str,
                 profile_dir: str, login_email: str = "",
                 on_password_saved=None, main_window: "MainWindow | None" = None,
                 parent=None):
        super().__init__(parent)
        self.setObjectName("gvSetupDialog")
        self.controller = controller
        self._main = main_window
        self._on_password_saved = on_password_saved
        self._login_email = login_email
        self._profile_dir = profile_dir
        self.setWindowTitle(f"Google Voice — {account_label}")
        self.setMinimumSize(960, 740)
        self.resize(980, 760)

        lay = QVBoxLayout(self)
        lay.setSpacing(14)
        lay.setContentsMargins(20, 18, 20, 18)

        lay.addWidget(_label(f"Connect {account_label}", "heroTitle"))
        sub = _label(
            "Sign in once below. Your session stays on this computer so you can "
            "dial without signing in again.",
            "muted",
        )
        sub.setWordWrap(True)
        lay.addWidget(sub)

        cred_row = QHBoxLayout()
        cred_row.addWidget(_label("Password", "muted"))
        self.e_password = QLineEdit()
        self.e_password.setEchoMode(QLineEdit.EchoMode.Password)
        self.e_password.setPlaceholderText("Google account password")
        if controller._login_password:
            self.e_password.setText(controller._login_password)
        cred_row.addWidget(self.e_password, stretch=1)
        apply_btn = _btn("Apply & sign in", "green")
        apply_btn.clicked.connect(self._apply_password)
        cred_row.addWidget(apply_btn)
        lay.addLayout(cred_row)

        if login_email:
            lay.addWidget(_label(f"Email: {login_email}", "accent"))

        self.load_bar = QProgressBar()
        self.load_bar.setRange(0, 0)
        self.load_bar.setTextVisible(False)
        self.load_bar.setFixedHeight(4)
        lay.addWidget(self.load_bar)

        self.lbl_status = _label("Preparing sign-in page…", "statusPill")
        self.lbl_status.setObjectName("statusPill")
        lay.addWidget(self.lbl_status)

        self.browser_frame = QFrame()
        self.browser_frame.setObjectName("browserFrame")
        self.browser_frame.setMinimumHeight(420)
        flay = QVBoxLayout(self.browser_frame)
        flay.setContentsMargins(0, 0, 0, 0)
        if main_window:
            main_window._embed_browser_visible(controller.view, flay)
        else:
            flay.addWidget(controller.view)
        lay.addWidget(self.browser_frame, stretch=1)

        btn_row = QHBoxLayout()
        reload_btn = _btn("Reload", "")
        reload_btn.clicked.connect(self._reload_signin)
        open_btn = _btn("Profile folder", "")
        open_btn.clicked.connect(self._open_profile)
        btn_row.addWidget(reload_btn)
        btn_row.addWidget(open_btn)
        btn_row.addStretch()
        lay.addLayout(btn_row)

        btn_done = _btn("I'm signed in — continue", "primary")
        btn_done.clicked.connect(self._confirm_login)
        lay.addWidget(btn_done)

        def on_log(_sid: int, msg: str) -> None:
            self.lbl_status.setText(msg)
            if "failed" not in msg.lower():
                self.load_bar.setRange(0, 1)
                self.load_bar.setValue(1)

        def on_login(_sid: int) -> None:
            self.controller.mark_logged_in()
            self.lbl_status.setText("Signed in — saving session…")
            self.login_succeeded.emit()
            QTimer.singleShot(700, self.accept)

        controller.log_message.connect(on_log)
        controller.login_detected.connect(on_login)
        controller._page.loadProgress.connect(self._on_load_progress)

        QTimer.singleShot(50, self._start_signin)

    def _start_signin(self) -> None:
        self.load_bar.setRange(0, 0)
        if self.controller._login_password:
            self.controller.load(for_setup=True)
        else:
            self.lbl_status.setText(
                "Enter password above, then click Apply & sign in.")

    def _reload_signin(self) -> None:
        self.load_bar.setRange(0, 0)
        self.lbl_status.setText("Reloading sign-in page…")
        self.controller.load(for_setup=True)

    def _open_profile(self) -> None:
        os.makedirs(self._profile_dir, exist_ok=True)
        os.startfile(self._profile_dir)

    def _on_load_progress(self, pct: int) -> None:
        if pct >= 100:
            self.load_bar.setRange(0, 1)
            self.load_bar.setValue(1)
            self.lbl_status.setText("Complete sign-in in the window below")

    def _apply_password(self) -> None:
        pw = self.e_password.text().strip()
        if not pw:
            self.lbl_status.setText("Enter your Google password first.")
            return
        if self._on_password_saved:
            self._on_password_saved(pw)
        self.controller.set_login_credentials(
            self._login_email or self.controller._login_email, pw)
        self.controller._autofill_paused = False
        self.controller._email_step_done = False
        self.controller._last_login_fill_status = ""
        self.load_bar.setRange(0, 0)
        self.lbl_status.setText("Signing in automatically…")
        self.controller.load(for_setup=True)

    def _confirm_login(self) -> None:
        self.controller.mark_logged_in()
        self.accept()

    def accept(self) -> None:
        if not self.controller.is_session_ready():
            self.controller.mark_logged_in()
        super().accept()


# ══════════════════════════════════════════════════════════════════════════════
#  ADD GOOGLE VOICE ACCOUNT DIALOG
# ══════════════════════════════════════════════════════════════════════════════

class AddGVAccountDialog(QDialog):
    """Single form: label, email, password — used for client-ready account setup."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Google Voice account")
        self.setMinimumWidth(440)
        lay = QVBoxLayout(self)
        lay.setSpacing(12)

        lay.addWidget(_label("Add a voice line", "heroTitle"))
        lay.addWidget(_label(
            "Password is stored only on this computer and used to sign in "
            "automatically in the background.",
            "muted",
        ))

        form = QFormLayout()
        form.setSpacing(10)
        self.e_name = QLineEdit()
        self.e_name.setPlaceholderText("e.g. Sales line 1")
        self.e_email = QLineEdit()
        self.e_email.setPlaceholderText("name@gmail.com")
        self.e_pw = QLineEdit()
        self.e_pw.setEchoMode(QLineEdit.EchoMode.Password)
        self.e_pw.setPlaceholderText("Google account password")
        self.e_notes = QLineEdit()
        self.e_notes.setPlaceholderText("Optional")
        form.addRow("Display name:", self.e_name)
        form.addRow("Google email:", self.e_email)
        form.addRow("Password:", self.e_pw)
        form.addRow("Notes:", self.e_notes)
        lay.addLayout(form)

        self.chk_auto = QComboBox()
        self.chk_auto.addItems([
            "Sign in automatically in background (recommended)",
            "I will connect manually later",
        ])
        lay.addWidget(self.chk_auto)

        row = QHBoxLayout()
        row.addStretch()
        cancel = _btn("Cancel", "secondary")
        cancel.clicked.connect(self.reject)
        ok = _btn("Add account", "primary")
        ok.clicked.connect(self._validate)
        row.addWidget(cancel)
        row.addWidget(ok)
        lay.addLayout(row)

    def _validate(self) -> None:
        if not self.e_name.text().strip() or not self.e_email.text().strip():
            QMessageBox.warning(self, "Required fields",
                                "Display name and Google email are required.")
            return
        if self.auto_login() and not self.e_pw.text().strip():
            QMessageBox.warning(self, "Password required",
                                "Enter the Google password for automatic sign-in.")
            return
        self.accept()

    def auto_login(self) -> bool:
        return self.chk_auto.currentIndex() == 0

    def account_data(self) -> dict:
        return {
            "name": self.e_name.text().strip(),
            "email": self.e_email.text().strip().lower(),
            "password": self.e_pw.text(),
            "notes": self.e_notes.text().strip(),
        }


# ══════════════════════════════════════════════════════════════════════════════
#  SLOT MONITOR (live line browser)
# ══════════════════════════════════════════════════════════════════════════════

class SlotMonitorDialog(QDialog):
    """Shows the Google Voice browser for one line so the user can talk on the call."""

    def __init__(self, controller: GVController, line_label: str,
                 main_window: "MainWindow", parent=None,
                 debug_monitor: bool = False):
        super().__init__(parent)
        self.controller = controller
        self._main = main_window
        self._released = False
        self._debug_monitor = debug_monitor
        title = "Monitor audio" if debug_monitor else "Live line"
        self.setWindowTitle(f"{title} - {line_label}")
        self.setWindowFlags(
            self.windowFlags()
            | Qt.WindowType.WindowMinimizeButtonHint
            | Qt.WindowType.WindowMaximizeButtonHint
        )
        self.setMinimumSize(960, 640)
        self.resize(1280, 820)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(10, 10, 10, 10)
        msg = (
            "Debug monitor: audio is enabled for this line so you can hear ringing, "
            "voicemail, or pickup while testing."
            if debug_monitor else
            "Use this Google Voice panel for the answered call. "
            "Your mic and speakers stay attached to this embedded browser."
        )
        lay.addWidget(_label(msg, "muted"))
        self.lbl_monitor = _label("Loading Google Voice view…", "statusPill")
        lay.addWidget(self.lbl_monitor)
        frame = QFrame()
        frame.setObjectName("browserFrame")
        frame.setMinimumHeight(520)
        fl = QVBoxLayout(frame)
        fl.setContentsMargins(0, 0, 0, 0)
        main_window._embed_browser_visible(controller.view, fl)
        self.controller.set_audio_muted(False)
        lay.addWidget(frame, stretch=1)
        btn_row = QHBoxLayout()
        refresh_btn = _btn("Redraw view", "secondary")
        refresh_btn.setToolTip(
            "Redraw the existing Google Voice screen without reloading the call.")
        refresh_btn.clicked.connect(self._refresh_view)
        btn_row.addWidget(refresh_btn)
        min_btn = _btn("Minimize", "secondary")
        min_btn.clicked.connect(self.showMinimized)
        btn_row.addWidget(min_btn)
        self._max_btn = _btn("Maximize", "secondary")
        self._max_btn.clicked.connect(self._toggle_maximized)
        btn_row.addWidget(self._max_btn)
        btn_row.addStretch()
        release_btn = _btn("Release slot / next", "green")
        release_btn.clicked.connect(self._release_slot)
        btn_row.addWidget(release_btn)
        end_btn = _btn("End call", "red")
        end_btn.clicked.connect(self._end_call)
        btn_row.addWidget(end_btn)
        lay.addLayout(btn_row)
        close_btn = _btn("Close monitor", "primary")
        close_btn.clicked.connect(self.accept)
        lay.addWidget(close_btn)
        controller.log_message.connect(self._on_log)

    def showEvent(self, event) -> None:
        super().showEvent(event)
        if not self.isMaximized():
            self._fit_to_screen()
        self.controller.prepare_for_visible_display()
        self.lbl_monitor.setText("Showing current Google Voice screen.")

    def _fit_to_screen(self) -> None:
        parent = self.parent()
        screen = QApplication.screenAt(parent.mapToGlobal(QPoint(0, 0))) if parent else None
        screen = screen or QApplication.primaryScreen()
        if screen is None:
            self.showMaximized()
            return
        rect = screen.availableGeometry()
        margin = 24
        self.setGeometry(
            rect.x() + margin,
            rect.y() + margin,
            max(self.minimumWidth(), rect.width() - margin * 2),
            max(self.minimumHeight(), rect.height() - margin * 2),
        )

    def _toggle_maximized(self) -> None:
        if self.isMaximized():
            self.showNormal()
            self._fit_to_screen()
            self._max_btn.setText("Maximize")
        else:
            self.showMaximized()
            self._max_btn.setText("Restore")
        QTimer.singleShot(80, self.controller.prepare_for_visible_display)

    def _on_log(self, _sid: int, msg: str) -> None:
        self.lbl_monitor.setText(msg)

    def _refresh_view(self) -> None:
        self.controller.prepare_for_visible_display()
        self.lbl_monitor.setText(
            "Redrew the existing Google Voice screen without reloading the call.")

    def _release_slot(self) -> None:
        self._main._next_call(self.controller.slot_id)
        self.accept()

    def _end_call(self) -> None:
        self._main._cut_call(self.controller.slot_id)
        self.accept()

    def _release_browser(self) -> None:
        if self._released:
            return
        self._released = True
        self.controller.set_audio_muted(True)
        try:
            self.controller.log_message.disconnect(self._on_log)
        except TypeError:
            pass
        self._main._hide_browser_after_setup(self.controller.view)

    def closeEvent(self, event) -> None:
        self._release_browser()
        super().closeEvent(event)

    def accept(self) -> None:
        self._release_browser()
        super().accept()

    def reject(self) -> None:
        self._release_browser()
        super().reject()


# ══════════════════════════════════════════════════════════════════════════════
#  CREATE USER DIALOG (client accounts — agent role only)
# ══════════════════════════════════════════════════════════════════════════════

class CreateUserDialog(QDialog):
    def __init__(self, db: CRMDatabase, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Add client user")
        self.setMinimumWidth(420)
        lay = QVBoxLayout(self)

        lay.addWidget(_label("New dialer user", "heroTitle"))
        lay.addWidget(_label(
            "Creates a standard user account for your client. "
            "They cannot change Google Voice settings or manage other users.",
            "muted",
        ))
        lay.addSpacing(8)

        form = QFormLayout()
        self.e_name = QLineEdit()
        self.e_name.setPlaceholderText("Client or agent name")
        self.e_email = QLineEdit()
        self.e_email.setPlaceholderText("client@company.com")
        self.e_pw = QLineEdit()
        self.e_pw.setEchoMode(QLineEdit.EchoMode.Password)
        self.e_pw.setPlaceholderText("At least 8 characters")
        self.plan = QComboBox()
        self.plan.addItems(["15 days", "Weekly", "Monthly", "No expiry"])
        self.max_slots = QSpinBox()
        self.max_slots.setRange(1, 5)
        self.max_slots.setValue(5)
        form.addRow("Name:", self.e_name)
        form.addRow("Email:", self.e_email)
        form.addRow("Password:", self.e_pw)
        form.addRow("Subscription:", self.plan)
        form.addRow("Max live slots:", self.max_slots)
        lay.addLayout(form)
        lay.addSpacing(12)

        btn = _btn("Create user", "primary")
        btn.clicked.connect(self._create)
        lay.addWidget(btn)

    def _create(self):
        name = self.e_name.text().strip()
        email = self.e_email.text().strip()
        pw = self.e_pw.text()
        if not all([name, email, pw]):
            QMessageBox.warning(self, "Missing fields", "All fields are required.")
            return
        if len(pw) < 8:
            QMessageBox.warning(self, "Password", "Use at least 8 characters.")
            return
        plan, expires_at = self._subscription_values()
        try:
            self.db.create_user(
                email, name, pw, role="agent",
                subscription_plan=plan,
                subscription_expires_at=expires_at,
                max_slots=self.max_slots.value(),
            )
            QMessageBox.information(
                self, "User created",
                f"{name} can sign in with:\n\nEmail: {email}\n"
                "Password: (the one you just set)\n\n"
                f"Subscription: {plan}\n"
                f"Expires: {expires_at or 'No local expiry'}\n"
                f"Max slots: {self.max_slots.value()}\n\n"
                "They will see Dialer, Live Calls, Logs, and CRM only.",
            )
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def _subscription_values(self) -> tuple[str, str]:
        label = self.plan.currentText()
        days = {
            "15 days": 15,
            "Weekly": 7,
            "Monthly": 30,
        }.get(label)
        if not days:
            return "No expiry", ""
        expires = datetime.now() + timedelta(days=days)
        return label, expires.strftime("%Y-%m-%d %H:%M:%S")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN APPLICATION WINDOW
# ══════════════════════════════════════════════════════════════════════════════

class MainWindow(QMainWindow):

    def __init__(self, db: CRMDatabase, user: dict, cfg: dict):
        super().__init__()
        self.db   = db
        self.user = user
        self.cfg  = cfg
        self._client_workstation = is_client_deployment(cfg)

        self.setWindowTitle("FT Solutions - Auto Dialer")
        self.setWindowIcon(_icon())
        self.resize(1280, 820)
        self.setMinimumSize(1024, 680)

        # ── Dialer state ──────────────────────────────────────────────────────
        self._controllers: list[GVController] = []
        self._contacts:    list[tuple[str, str]] = []
        self._contact_idx: int = 0
        self._running:     bool = False
        self._slot_start:  dict[int, float] = {}
        self._slot_phone:  dict[int, str]   = {}
        self._slot_name:   dict[int, str]   = {}
        self._slot_retry_attempt: dict[int, int] = {}
        self._slot_detection_debug: dict[int, dict] = {}
        self._slot_cooldown_until: dict[int, float] = {}
        self._slot_ui_state: dict[int, str] = {}
        self._manual_test_slots: set[int] = set()
        self._slot_call_token: dict[int, int] = {}
        self._slot_final_logged: set[int] = set()
        self._slot_call_times: dict[int, dict[str, str]] = {}
        self._recovering_slots: set[int] = set()
        self._controllers_ready: bool = False
        self._controllers_boot_remaining: int = 0
        self._effective_slots: int = 1
        self._campaign_generation: int = 0
        self._all_logs:    list = []
        self._retry_queue = DialRetryQueue(
            max_retries=int(cfg.get("max_retries", 3)),
            backoff_sec=tuple(cfg.get("retry_backoff_sec", [5, 15, 45])),
        )
        self._watchdog = SlotWatchdog(self)
        self._watchdog.slot_restart_requested.connect(self._restart_slot)
        self._slot_restart_cooldown: dict[int, float] = {}
        self._pending_slot_restarts: dict[int, dict] = {}
        self._slot_monitors: dict[int, SlotMonitorDialog] = {}
        self._configure_watchdog()
        self._gv_accounts: list[dict] = load_gv_accounts()
        self._headless_login_queue: list[dict] = []
        self._pacing_engine = PredictivePacingEngine(
            PacingConfig(
                max_dials_per_interval=int(cfg.get("max_concurrent_dials", 3)),
                pacing_interval_sec=float(cfg.get("pacing_interval_sec", 2.5)),
            )
        )
        self._pacing_dial_history: list[int] = []
        self._slot_agent_handled: set[int] = set()
        self._ws_thread = None
        self._dial_assign_count: int = 0
        self._dial_started_at: float = 0.0

        # ── Timers ────────────────────────────────────────────────────────────
        self._dial_timer   = QTimer(self)    # fires to assign next number to free slot
        self._dial_timer.setInterval(2500)
        self._dial_timer.timeout.connect(self._assign_pending_calls)

        self._assign_debounce = QTimer(self)
        self._assign_debounce.setSingleShot(True)
        self._assign_debounce.timeout.connect(self._assign_pending_calls)

        self._elapsed_timer = QTimer(self)   # updates elapsed display on slot cards
        self._elapsed_timer.setInterval(1000)
        self._elapsed_timer.timeout.connect(self._tick_elapsed)

        self._headless_timer = QTimer(self)
        self._headless_timer.setInterval(2000)
        self._headless_timer.timeout.connect(self._tick_headless_logins)

        self._stuck_dial_timer = QTimer(self)
        self._stuck_dial_timer.setInterval(10000)
        self._stuck_dial_timer.timeout.connect(self._check_stuck_dials)

        self._health_timer = QTimer(self)
        self._health_timer.setInterval(5000)
        self._health_timer.timeout.connect(self._update_health_panel)

        # ── Build UI ──────────────────────────────────────────────────────────
        self._build_hidden_browser_container()
        self._build_header()
        self._build_tabs()
        self._build_status_bar()

        self.tabs.currentChanged.connect(self._on_tab_changed)

        # ── Boot controllers (deferred — UI shows first) ─────────────────────
        requested = int(cfg.get("n_slots", 1))
        self._effective_slots = recommended_slots(requested)
        self.statusBar().showMessage("Starting — Google Voice loads in a few seconds…")
        QTimer.singleShot(1800, self._deferred_boot_controllers)
        self._health_timer.start()
        log_info(f"Session started — user {user.get('email', '?')}")

    def closeEvent(self, event) -> None:
        self._watchdog.stop()
        self._dial_timer.stop()
        self._assign_debounce.stop()
        self._elapsed_timer.stop()
        self._headless_timer.stop()
        self._stuck_dial_timer.stop()
        self._health_timer.stop()
        self._save_campaign_progress()
        for dialog in list(self._slot_monitors.values()):
            try:
                dialog.close()
            except Exception:
                pass
        self._slot_monitors.clear()
        for ctrl in list(self._controllers):
            if ctrl is not None:
                self._dispose_controller(ctrl)
        self._controllers.clear()
        QApplication.processEvents()
        super().closeEvent(event)

    def _deferred_boot_controllers(self) -> None:
        """Load Google Voice after the UI is visible (prevents startup freeze)."""
        requested = int(self.cfg.get("n_slots", 1))
        self._effective_slots = recommended_slots(requested)
        reason = low_resource_reason(requested, self._effective_slots)
        if reason:
            log_warning(reason)
            self._log(reason)
            if hasattr(self, "spin_slots"):
                self.spin_slots.setValue(self._effective_slots)
            QMessageBox.information(
                self,
                "Low-resource mode",
                f"{reason}\n\n"
                "You do NOT need a new PC — close extra Chrome windows "
                "and use 1 line at a time on 8 GB RAM.\n\n"
                f"Chrome tabs/processes now: {chrome_process_count()}\n"
                f"System RAM: ~{system_ram_gb():.0f} GB",
            )
        self._controllers_boot_remaining = self._effective_slots
        self._init_controllers(self._effective_slots)

    def _on_controller_booted(self) -> None:
        self._controllers_boot_remaining = max(0, self._controllers_boot_remaining - 1)
        if self._controllers_boot_remaining == 0 and not self._controllers_ready:
            self._controllers_ready = True
            if not self._watchdog._running:
                self._watchdog.start()
            self._update_health_panel()
            self.statusBar().showMessage("Google Voice lines ready")
            self._refresh_slot_login_badges()
            log_info("All voice line browsers loaded")

    def _gv_runtime_cfg(self) -> dict:
        audio_on = effective_enable_ai_audio(self.cfg)
        return {
            "call_timeout": int(self.cfg.get("call_timeout", 60)),
            "enable_ai_audio": audio_on,
            "amd_mode": str(self.cfg.get("amd_mode", "heuristic") or "heuristic"),
            "amd_early_decision_ms": int(self.cfg.get("amd_early_decision_ms", 800)),
            "amd_max_decision_ms": int(self.cfg.get("amd_max_decision_ms", 2500)),
            "amd_human_first_seconds": float(self.cfg.get("amd_human_first_seconds", 5)),
            "audio_device": self.cfg.get("audio_device", ""),
            "live_debug_mode": bool(self.cfg.get("live_debug_mode", False)),
        }

    def _apply_runtime_cfg_to_controllers(self) -> None:
        runtime = self._gv_runtime_cfg()
        for ctrl in self._controllers:
            if ctrl is not None:
                ctrl.apply_runtime_cfg(runtime)
                if effective_enable_ai_audio(runtime):
                    ctrl._audio_monitor.prime()

    def _configure_watchdog(self) -> None:
        stuck = float(self.cfg.get(
            "watchdog_stuck_state_sec",
            max(90, self.cfg.get("call_timeout", 60) + 30),
        ))
        self._watchdog.configure(
            heartbeat_timeout_sec=float(
                self.cfg.get("watchdog_heartbeat_timeout_sec", 45)),
            stuck_state_sec=stuck,
            memory_limit_mb=int(self.cfg.get("slot_memory_limit_mb", 1200)),
            recycle_after_calls=int(self.cfg.get("slot_recycle_after_calls", 75)),
            check_interval_ms=int(
                float(self.cfg.get("watchdog_check_interval_sec", 5)) * 1000),
        )
        self._watchdog.set_memory_getter(webengine_total_memory_mb)

    # ── Hidden browser container ──────────────────────────────────────────────

    def _build_hidden_browser_container(self):
        """
        All QWebEngineViews live here — hidden from the agent.
        Size 1×1 so they are technically in the layout (needed for WebRTC audio)
        but invisible to the user.
        """
        self._browser_host = QWidget(self)
        self._browser_host.setMaximumSize(1, 1)
        self._browser_layout = QHBoxLayout(self._browser_host)
        self._browser_layout.setContentsMargins(0, 0, 0, 0)

    def _show_browser_for_setup(self, view: QWebEngineView) -> None:
        """Expand embedded browser for visible login in GVSetupDialog."""
        if view.parent() is self._browser_host:
            self._browser_layout.removeWidget(view)
            view.setParent(None)
        max_dim = 16777215
        view.setMinimumSize(400, 300)
        view.setMaximumSize(max_dim, max_dim)
        view.setWindowFlag(Qt.WindowType.Tool, False)
        view.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        view.setAttribute(Qt.WidgetAttribute.WA_DontShowOnScreen, False)
        view.show()
        view.updateGeometry()
        view.repaint()

    def _embed_browser_visible(self, view: QWebEngineView,
                               layout: QVBoxLayout) -> GVController | None:
        """Place a slot browser in a visible dialog layout (Listen / setup)."""
        ctrl: GVController | None = None
        for c in self._controllers:
            if c is None:
                continue
            if c.view is view:
                ctrl = c
                break
        if view.parent() is self._browser_host:
            self._browser_layout.removeWidget(view)
        view.setParent(None)
        layout.addWidget(view, stretch=1)
        self._show_browser_for_setup(view)
        if ctrl:
            ctrl.prepare_for_visible_display()
        return ctrl

    def _hide_browser_after_setup(self, view: QWebEngineView) -> None:
        """Return embedded browser to hidden 1×1 host after login setup."""
        for ctrl in self._controllers:
            if ctrl and ctrl.view is view:
                ctrl.prepare_for_background_rendering()
                return
        view.setParent(None)
        view.setAttribute(Qt.WidgetAttribute.WA_DontShowOnScreen, True)
        view.resize(1280, 900)
        view.show()

    def _dispose_controller(self, ctrl: GVController) -> None:
        """Remove a slot browser from the layout and tear down WebEngine safely."""
        view = getattr(ctrl, "view", None)
        if view is not None:
            if view.parent() is self._browser_host:
                self._browser_layout.removeWidget(view)
            elif view.parent() is not None:
                view.setParent(None)
        ctrl.shutdown()
        ctrl.deleteLater()
        QApplication.processEvents()

    def _refresh_slot_login_badges(self) -> None:
        if not hasattr(self, "_slot_cards"):
            return
        for ctrl in self._controllers:
            if ctrl is None:
                continue
            sid = ctrl.slot_id
            if sid in self._slot_cards:
                self._slot_cards[sid].set_gv_login_ready(ctrl.is_session_ready())

    def _controller_for_profile(self, profile_name: str) -> GVController | None:
        target = gv_profile_dir(profile_name)
        for ctrl in self._controllers:
            if ctrl is None:
                continue
            if os.path.abspath(ctrl.profile_dir) == os.path.abspath(target):
                return ctrl
        return None

    def _account_session_ready(self, acct: dict) -> bool:
        target = gv_profile_dir(acct["profile"])
        if gv_has_session_marker(target):
            return True
        ctrl = self._controller_for_profile(acct["profile"])
        if ctrl and ctrl.is_session_ready():
            return True
        if ctrl:
            ctrl._check_login()
        return gv_has_session_marker(target)

    def _ensure_profile_controller(self, acct: dict) -> GVController | None:
        ctrl = self._controller_for_profile(acct["profile"])
        if ctrl:
            return ctrl
        slot_id = len(self._controllers)
        ctrl = GVController(
            slot_id,
            gv_profile_dir(acct["profile"]),
            parent=self,
            profile_key=acct["profile"],
            login_email=acct.get("email", ""),
            login_password=acct.get("password", ""),
            runtime_cfg=self._gv_runtime_cfg(),
        )
        ctrl.state_changed.connect(self._on_slot_state)
        ctrl.login_detected.connect(self._on_slot_login)
        ctrl.log_message.connect(self._on_slot_log)
        ctrl.detection_update.connect(self._on_slot_detection)
        ctrl.prepare_for_background_rendering()
        self._controllers.append(ctrl)
        if gv_has_session_marker(ctrl.profile_dir):
            ctrl.mark_logged_in()
        return ctrl

    def _start_headless_login(self, acct: dict) -> None:
        """Background Google sign-in using saved password (no setup window)."""
        if not acct.get("password"):
            return
        ctrl = self._ensure_profile_controller(acct)
        if not ctrl:
            return
        ctrl.set_login_credentials(acct.get("email", ""), acct.get("password", ""))
        if ctrl.is_session_ready():
            self._refresh_slot_login_badges()
            return
        ctrl.load()
        self._headless_login_queue = [
            j for j in self._headless_login_queue
            if j.get("profile") != acct["profile"]
        ]
        self._headless_login_queue.append({
            "profile": acct["profile"],
            "name": acct.get("name", ""),
            "attempts": 0,
        })
        self._headless_timer.start()
        self._log(f"Signing in {acct.get('name', 'account')} in the background…")

    def _tick_headless_logins(self) -> None:
        pending: list[dict] = []
        for job in self._headless_login_queue:
            ctrl = self._controller_for_profile(job["profile"])
            job["attempts"] = job.get("attempts", 0) + 1
            target = gv_profile_dir(job["profile"])
            if gv_has_session_marker(target) or (ctrl and ctrl.is_session_ready()):
                if ctrl:
                    ctrl.mark_logged_in()
                self._log(f"{job.get('name', 'Account')} is ready to dial")
                continue
            if job["attempts"] > 45:
                self._log(
                    f"{job.get('name', 'Account')}: automatic sign-in incomplete — "
                    "use Settings → Connect account")
                continue
            if ctrl:
                ctrl._check_login()
                if job["attempts"] % 3 == 0:
                    ctrl._try_auto_login()
            pending.append(job)
        self._headless_login_queue = pending
        if not pending:
            self._headless_timer.stop()
        self._refresh_slot_login_badges()

    def _open_slot_monitor(self, slot_id: int, debug_monitor: bool = False) -> None:
        self._slot_agent_handled.add(slot_id)
        ctrl = self._get_ctrl(slot_id)
        if not ctrl:
            QMessageBox.information(
                self, "No line",
                "This dialing line is not active yet.")
            return
        label = self._slot_label(slot_id)
        existing = self._slot_monitors.get(slot_id)
        if existing and existing.isVisible():
            if existing.isMinimized():
                existing.showNormal()
            existing.raise_()
            existing.activateWindow()
            return
        dlg = SlotMonitorDialog(
            ctrl, label, self, self, debug_monitor=debug_monitor)
        dlg.finished.connect(lambda _=0, sid=slot_id: self._slot_monitors.pop(sid, None))
        self._slot_monitors[slot_id] = dlg
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()

    def _focus_answered_slot(self, slot_id: int) -> None:
        self.tabs.setCurrentWidget(self.tab_live)
        if hasattr(self, "_slot_cards") and slot_id in self._slot_cards:
            card = self._slot_cards[slot_id]
            card.raise_()
            card.setFocus()

    def _open_slot_audio_monitor(self, slot_id: int) -> None:
        self._open_slot_monitor(slot_id, debug_monitor=True)

    def _test_call(self, slot_id: int) -> None:
        ctrl = self._get_ctrl(slot_id)
        acct = self._slot_account(slot_id)
        if not ctrl or not acct:
            QMessageBox.warning(
                self, "Line not configured",
                f"Add a Google Voice account for Line {slot_id + 1} first.")
            return
        if not self._account_session_ready(acct):
            QMessageBox.warning(
                self, "Google Voice Not Ready",
                "Open Settings -> Connect account and complete sign-in for "
                f"{acct.get('name') or acct.get('email')}.")
            return
        if self._slot_has_active_phone(slot_id) or ctrl.current_state not in ("IDLE", "ENDED", "ENDED_MANUALLY", "BUSY", "FAILED"):
            QMessageBox.information(
                self, "Line busy",
                f"Line {slot_id + 1} already has an active call.")
            return
        from PyQt6.QtWidgets import QInputDialog
        raw, ok = QInputDialog.getText(
            self, "Test call", f"Phone number for Line {slot_id + 1}:")
        if not ok:
            return
        d10 = clean_phone(raw)
        if not d10:
            QMessageBox.warning(self, "Invalid number", "Enter a valid phone number.")
            return
        phone = fmt_e164(d10)
        self._log(f"[Slot {slot_id}] Test call requested for {phone}")
        self._manual_test_slots.add(slot_id)
        self._dial_on_slot(ctrl, phone, "Test call", 0)

    def _dialing_login_ok(self) -> tuple[bool, str]:
        if not self._gv_accounts:
            return False, (
                "Add at least one Google Voice account in Settings, then "
                "use Connect account.")
        n = self.spin_slots.value()
        if self.user.get("role") != "admin":
            max_slots = int(self.user.get("max_slots") or 1)
            if n > max_slots:
                return False, (
                    f"Your subscription allows {max_slots} live slot(s). "
                    "Reduce Lines at once or contact your administrator.")
        if len(self._gv_accounts) < n:
            return False, (
                f"You selected {n} live slot(s), but only "
                f"{len(self._gv_accounts)} Google Voice account(s) are configured.\n\n"
                "Add one signed-in Google Voice account for each slot. "
                "For 5-slot dialing, configure 5 ready lines.")
        distinct_lines = {
            str(acct.get("email") or acct.get("profile") or "").strip().lower()
            for acct in self._gv_accounts[:n]
            if str(acct.get("email") or acct.get("profile") or "").strip()
        }
        if len(distinct_lines) < n:
            return False, (
                f"You selected {n} live slot(s), but only {len(distinct_lines)} "
                "distinct Google Voice line(s) are configured.\n\n"
                "Each realtime slot needs its own signed-in Google Voice account. "
                "Duplicate browser profiles for the same email cannot dial concurrently.")
        missing: list[str] = []
        for i in range(n):
            acct = self._slot_account(i)
            if not acct:
                missing.append(f"Slot {i + 1}")
                continue
            if not self._account_session_ready(acct):
                missing.append(acct.get("name") or acct.get("email", f"Slot {i+1}"))
        if missing:
            return False, (
                "Google Voice is not ready for:\n• "
                + "\n• ".join(missing)
                + "\n\nOpen Settings → Connect account and complete sign-in once.")
        return True, ""

    # ── Header ────────────────────────────────────────────────────────────────

    def _build_header(self):
        hdr = QWidget()
        hdr.setObjectName("appHeader")
        hdr.setFixedHeight(76)
        h = QHBoxLayout(hdr)
        h.setContentsMargins(20, 10, 20, 10)

        # Logo + name
        left = QHBoxLayout()
        px = _pixmap(52)
        if px:
            lbl_logo = QLabel()
            lbl_logo.setPixmap(px)
            lbl_logo.setStyleSheet("background: transparent;")
            left.addWidget(lbl_logo)
            left.addSpacing(12)
        col = QVBoxLayout()
        col.setSpacing(2)
        name_lbl = _label("FT Solutions", bold=True, size=14)
        name_lbl.setObjectName("brandName")
        sub_lbl = _label("Auto Dialer Pro", "brandTagline")
        sub_lbl.setObjectName("brandTagline")
        col.addWidget(name_lbl)
        col.addWidget(sub_lbl)
        left.addLayout(col)
        h.addLayout(left)
        h.addStretch()

        # Right side
        right = QHBoxLayout()
        right.setSpacing(10)

        # User badge
        role_name = "Administrator" if self.user["role"] == "admin" else "Agent"
        user_lbl = QLabel(
            f'<span style="color:#1e293b;"><b>{self.user["name"]}</b></span>'
            f' &nbsp;·&nbsp; <span style="color:#64748b;">{role_name}</span>'
        )
        user_lbl.setObjectName("headerUser")
        right.addWidget(user_lbl)

        div = QFrame()
        div.setFrameShape(QFrame.Shape.VLine)
        right.addWidget(div)

        wa_btn = _btn("Support", "wa")
        import webbrowser as _wb
        wa_btn.setToolTip(WA_NUMBER)
        wa_btn.clicked.connect(lambda: _wb.open(WHATSAPP_URL))
        right.addWidget(wa_btn)

        div2 = QFrame()
        div2.setFrameShape(QFrame.Shape.VLine)
        right.addWidget(div2)

        self._theme_btn = _btn("Dark mode", "ghost")
        self._theme_btn.setMinimumWidth(100)
        self._theme_btn.clicked.connect(self._toggle_theme)
        if self.cfg.get("theme", DEFAULT_THEME) == "light":
            self._theme_btn.setText("Dark mode")
        else:
            self._theme_btn.setText("Light mode")
        right.addWidget(self._theme_btn)

        logout_btn = _btn("Sign out", "ghost")
        logout_btn.clicked.connect(self._logout)
        right.addWidget(logout_btn)

        h.addLayout(right)

        # Central widget holds header + tabs
        central = QWidget()
        self.setCentralWidget(central)
        vlay = QVBoxLayout(central)
        vlay.setContentsMargins(0, 0, 0, 0)
        vlay.setSpacing(0)
        vlay.addWidget(self._browser_host)  # 1×1 hidden browsers
        vlay.addWidget(hdr)
        self._main_vlay = vlay  # tabs will be added next

    # ── Tabs ──────────────────────────────────────────────────────────────────

    def _build_tabs(self):
        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)

        self.tab_dialer = QWidget()
        self.tab_live   = QWidget()
        self.tab_logs   = QWidget()
        self.tab_crm    = QWidget()
        self.tab_settings = QWidget()

        self.tabs.addTab(self.tab_dialer,   "  Dialer  ")
        self.tabs.addTab(self.tab_live,     "  Live Calls  ")
        self.tabs.addTab(self.tab_logs,     "  Call Logs  ")
        self.tabs.addTab(self.tab_crm,      "  CRM  ")
        self.tabs.addTab(self.tab_settings, "  Settings  ")

        if self.user["role"] == "admin" and not self._client_workstation:
            self.tab_admin = QWidget()
            self.tabs.addTab(self.tab_admin, "  Administration  ")
            self._build_admin_tab()

        self._main_vlay.addWidget(self.tabs)

        self._build_dialer_tab()
        self._build_live_tab()
        self._build_logs_tab()
        self._build_crm_tab()
        self._build_settings_tab()

    def _apply_stable_mode(self) -> None:
        """Recommended settings for smooth dialing on most PCs."""
        slots = 1 if system_ram_gb() < 12 else 2
        self.spin_slots.setValue(slots)
        self.spin_timeout.setValue(60)
        self.spin_cooldown.setValue(6.0)
        self.spin_vm_hangup.setValue(4)
        self.cfg.update({
            "n_slots": slots,
            "call_timeout": 60,
            "cooldown": 6.0,
            "voicemail_hangup_sec": 4,
            "enable_ai_audio": False,
        })
        _save_cfg(self.cfg)
        self._effective_slots = recommended_slots(slots)
        self._apply_runtime_cfg_to_controllers()
        self._log(
            f"Stable mode applied — {self._effective_slots} line(s), "
            "60s timeout, 6s cooldown, AI audio off")
        QMessageBox.information(
            self,
            "Stable mode",
            f"Dialing options set for stability on your PC (~{system_ram_gb():.0f} GB RAM):\n\n"
            f"• {self._effective_slots} line(s) at once\n"
            "• 60 second call timeout\n"
            "• 6 second cooldown between calls\n"
            "• 4 second voicemail hangup\n"
            "• AI audio detection off\n\n"
            "Close extra Chrome windows before dialing.\n"
            "Restart the app to apply line count.",
        )

    def _update_health_panel(self) -> None:
        if not hasattr(self, "lbl_health_status"):
            return
        mb = webengine_total_memory_mb()
        limit = int(self.cfg.get("slot_memory_limit_mb", 700))
        recovering = len(self._recovering_slots) + len(self._pending_slot_restarts)
        active = sum(
            1 for c in self._controllers
            if c and c.current_state not in (
                "IDLE", "ENDED", "ENDED_MANUALLY", "FAILED", "NO_ANSWER", "BUSY")
        )
        parts = [f"WebEngine ~{mb} MB (limit {limit} MB)", f"Active lines: {active}"]
        if recovering:
            parts.append(f"Recovering {recovering} line(s)")
        if mb > limit:
            parts.append("High memory — consider fewer lines or restart")
        self.lbl_health_status.setText(" · ".join(parts))
        lp = log_path()
        if lp and hasattr(self, "lbl_health_log"):
            self.lbl_health_log.setText(f"Operations log: {lp}")

    def _mark_call_time(self, slot_id: int, field: str) -> None:
        times = self._slot_call_times.setdefault(slot_id, {})
        if field not in times:
            times[field] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def _build_status_bar(self):
        self.statusBar().showMessage("Ready")

    # ══════════════════════════════════════════════════════════════════════════
    #  DIALER TAB
    # ══════════════════════════════════════════════════════════════════════════

    def _build_dialer_tab(self):
        lay = QVBoxLayout(self.tab_dialer)
        lay.setSpacing(12)
        lay.setContentsMargins(16, 14, 16, 14)

        # File picker
        grp_file = QGroupBox("Contact list")
        flay = QHBoxLayout(grp_file)
        self.excel_input = QLineEdit(self.cfg.get("excel_path", ""))
        self.excel_input.setPlaceholderText("Choose an Excel file with phone numbers")
        self.excel_input.setReadOnly(True)
        flay.addWidget(self.excel_input)
        browse_btn = _btn("Browse…", "secondary")
        browse_btn.clicked.connect(self._browse)
        flay.addWidget(browse_btn)
        load_btn = _btn("Load contacts", "green")
        load_btn.clicked.connect(self._load_numbers)
        flay.addWidget(load_btn)
        crm_btn = _btn("Load CRM contacts", "secondary")
        crm_btn.setToolTip("Use saved CRM contacts without reloading Excel")
        crm_btn.clicked.connect(self._load_crm_numbers)
        flay.addWidget(crm_btn)
        test_btn = _btn("Sample list", "secondary")
        test_btn.setToolTip("Load the built-in test contact list")
        test_btn.clicked.connect(self._load_test_numbers)
        flay.addWidget(test_btn)
        lay.addWidget(grp_file)

        grp_settings = QGroupBox("Dialing options")
        slay = QHBoxLayout(grp_settings)
        slay.addWidget(_label("Lines at once:", "muted"))
        self.spin_slots = QSpinBox()
        self.spin_slots.setRange(1, 5)
        self.spin_slots.setValue(self.cfg.get("n_slots", 1))
        slay.addWidget(self.spin_slots)
        slay.addSpacing(20)
        slay.addWidget(_label("Call Timeout (sec):", "muted"))
        self.spin_timeout = QSpinBox()
        self.spin_timeout.setRange(20, 180)
        self.spin_timeout.setValue(self.cfg.get("call_timeout", 60))
        slay.addWidget(self.spin_timeout)
        slay.addSpacing(20)
        slay.addWidget(_label("Cooldown between calls (sec):", "muted"))
        self.spin_cooldown = QDoubleSpinBox()
        self.spin_cooldown.setRange(1.0, 30)
        self.spin_cooldown.setSingleStep(0.5)
        self.spin_cooldown.setValue(float(self.cfg.get("cooldown", 4.0)))
        self.spin_cooldown.setToolTip(
            "Pause on each line after a call ends before dialing the next number")
        slay.addWidget(self.spin_cooldown)
        slay.addSpacing(20)
        slay.addWidget(_label("Voicemail hangup (sec):", "muted"))
        self.spin_vm_hangup = QSpinBox()
        self.spin_vm_hangup.setRange(1, 15)
        self.spin_vm_hangup.setValue(int(self.cfg.get("voicemail_hangup_sec", 3)))
        slay.addWidget(self.spin_vm_hangup)
        slay.addStretch()
        lay.addWidget(grp_settings)

        # Progress
        grp_prog = QGroupBox("Campaign progress")
        play = QVBoxLayout(grp_prog)
        stat_row = QHBoxLayout()
        self.lbl_total   = _label("Total: —",     "statValue", bold=True)
        self.lbl_done    = _label("Completed: —", "statValue", bold=True)
        self.lbl_rem     = _label("Remaining: —", "statValue", bold=True)
        self.lbl_invalid = _label("Invalid: —",   "statValue", bold=True)
        for w in (self.lbl_total, self.lbl_done, self.lbl_rem, self.lbl_invalid):
            stat_row.addWidget(w)
            stat_row.addSpacing(20)
        stat_row.addStretch()
        play.addLayout(stat_row)
        self.progress = QProgressBar()
        self.progress.setValue(0)
        play.addWidget(self.progress)
        lay.addWidget(grp_prog)

        # Control buttons
        btn_row = QHBoxLayout()
        self.btn_start = _btn("Start dialing", "primary")
        self.btn_start.setEnabled(False)
        self.btn_start.setMinimumHeight(44)
        self.btn_start.clicked.connect(self._start_dialing)
        self.btn_stop  = _btn("Stop", "red")
        self.btn_stop.setMinimumHeight(44)
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self._stop_dialing)
        self.btn_emergency_stop = _btn("Emergency Stop", "red")
        self.btn_emergency_stop.setMinimumHeight(44)
        self.btn_emergency_stop.setEnabled(False)
        self.btn_emergency_stop.clicked.connect(self._emergency_stop)
        btn_row.addWidget(self.btn_start)
        btn_row.addWidget(self.btn_stop)
        btn_row.addWidget(self.btn_emergency_stop)
        btn_row.addStretch()
        lay.addLayout(btn_row)

        # Activity log
        grp_log = QGroupBox("Activity")
        llay = QVBoxLayout(grp_log)
        self.console = QTextEdit()
        self.console.setObjectName("console")
        self.console.setReadOnly(True)
        self.console.setMinimumHeight(120)
        self.console.setMaximumHeight(180)
        llay.addWidget(self.console)
        lay.addWidget(grp_log, stretch=1)

    # ══════════════════════════════════════════════════════════════════════════
    #  LIVE CALLS TAB
    # ══════════════════════════════════════════════════════════════════════════

    def _build_live_tab(self):
        lay = QVBoxLayout(self.tab_live)
        lay.setSpacing(12)
        lay.setContentsMargins(16, 14, 16, 14)

        info = _label(
            "Your calls run through Google Voice in the background. "
            "Each line shows live status. When someone answers, the card highlights "
            "and you can talk, then choose Next number or End call.",
            "muted"
        )
        info.setWordWrap(True)
        lay.addWidget(info)
        lay.addWidget(_hline())

        # Slot cards — grid (max 3 per row) inside scroll area for many lines
        self._cards_widget = QWidget()
        self._cards_layout = QGridLayout(self._cards_widget)
        self._cards_layout.setSpacing(14)
        self._cards_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setWidget(self._cards_widget)

        self._rebuild_slot_cards(self.cfg.get("n_slots", 5))
        lay.addWidget(scroll, stretch=1)

        # Bottom controls
        lay.addWidget(_hline())
        brow = QHBoxLayout()
        btn_start2 = _btn("Start dialing", "primary")
        btn_start2.clicked.connect(self._start_dialing)
        btn_stop2  = _btn("Stop", "red")
        btn_stop2.clicked.connect(self._stop_dialing)
        btn_emergency2 = _btn("Emergency Stop", "red")
        btn_emergency2.clicked.connect(self._emergency_stop)
        self.btn_start_live = btn_start2
        self.btn_stop_live = btn_stop2
        self.btn_emergency_live = btn_emergency2
        brow.addWidget(btn_start2)
        brow.addWidget(btn_stop2)
        brow.addWidget(btn_emergency2)
        brow.addStretch()
        lay.addLayout(brow)

    def _rebuild_slot_cards(self, n: int):
        # Clear existing
        while self._cards_layout.count():
            item = self._cards_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._slot_cards: dict[int, SlotCard] = {}
        cols = min(max(n, 1), 3)
        for i in range(n):
            card = SlotCard(i)
            card.set_line_label(self._slot_label(i))
            card.set_ai_audio_enabled(bool(self.cfg.get("enable_ai_audio", True)))
            card.next_clicked.connect(self._next_call)
            card.cut_clicked.connect(self._cut_call)
            card.listen_clicked.connect(self._open_slot_monitor)
            card.monitor_clicked.connect(self._open_slot_audio_monitor)
            card.test_clicked.connect(self._test_call)
            card.redial_clicked.connect(self._retry_slot_dial)
            row, col = divmod(i, cols)
            self._cards_layout.addWidget(card, row, col)
            self._slot_cards[i] = card
        # Balance column widths in the last row
        for c in range(cols):
            self._cards_layout.setColumnStretch(c, 1)

    # ══════════════════════════════════════════════════════════════════════════
    #  CALL LOGS TAB
    # ══════════════════════════════════════════════════════════════════════════

    def _build_logs_tab(self):
        lay = QVBoxLayout(self.tab_logs)
        lay.setSpacing(8)
        lay.setContentsMargins(16, 14, 16, 14)

        # Top row
        top = QHBoxLayout()
        top.addWidget(_label("Call History", bold=True, size=12))
        top.addStretch()
        for txt, fn, nm in [
            ("📤  Export", self._export_logs, "green"),
            ("🗑  Clear",   self._clear_logs,  "red"),
            ("🔄  Refresh", self._refresh_logs, ""),
        ]:
            b = _btn(txt, nm); b.clicked.connect(fn); top.addWidget(b)
        lay.addLayout(top)

        # Stat labels
        stat_row = QHBoxLayout()
        self.log_total = _label("Total: 0")
        self.log_ended = _label("Ended: 0")
        self.log_vm    = _label("Voicemail: 0")
        self.log_fail  = _label("Failed: 0")
        for w in (self.log_total, self.log_ended, self.log_vm, self.log_fail):
            stat_row.addWidget(w); stat_row.addSpacing(16)
        stat_row.addStretch()
        lay.addLayout(stat_row)

        # Filter bar
        frow = QHBoxLayout()
        frow.addWidget(QLabel("🔍 Filter:"))
        self.log_search = QLineEdit()
        self.log_search.setPlaceholderText("phone, status, date…")
        self.log_search.setMaximumWidth(240)
        self.log_search.textChanged.connect(self._apply_log_filter)
        frow.addWidget(self.log_search)
        frow.addSpacing(10)
        self.log_status_combo = QComboBox()
        self.log_status_combo.addItems(
            ["All Statuses", "ENDED", "ENDED_MANUALLY", "VOICEMAIL", "NO_ANSWER", "BUSY", "FAILED"])
        self.log_status_combo.currentTextChanged.connect(self._apply_log_filter)
        frow.addWidget(self.log_status_combo)
        frow.addStretch()
        lay.addLayout(frow)

        # Table
        self.log_table = QTableWidget(0, 7)
        self.log_table.setHorizontalHeaderLabels(
            ["Logged", "Phone", "Status", "Duration", "Line", "Dialed", "Connected"])
        self.log_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.log_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.log_table.setAlternatingRowColors(True)
        self.log_table.verticalHeader().setVisible(False)
        _configure_data_table(self.log_table, min_visible_rows=8)
        lay.addWidget(self.log_table, stretch=1)
        self._refresh_logs()
        self._refresh_logs()

    # ══════════════════════════════════════════════════════════════════════════
    #  CRM TAB
    # ══════════════════════════════════════════════════════════════════════════

    def _build_crm_tab(self):
        lay = QVBoxLayout(self.tab_crm)
        lay.setSpacing(8)
        lay.setContentsMargins(16, 14, 16, 14)

        top = QHBoxLayout()
        top.addWidget(_label("Contacts", bold=True, size=12))
        top.addStretch()
        for txt, fn, nm in [
            ("📥  Import Excel", self._import_contacts, ""),
            ("+ Add",           self._add_contact,      "green"),
            ("🗑  Delete",       self._delete_contact,   "red"),
            ("🔄  Refresh",      self._refresh_crm,      ""),
        ]:
            b = _btn(txt, nm); b.clicked.connect(fn); top.addWidget(b)
        lay.addLayout(top)

        # Status filter
        frow = QHBoxLayout()
        frow.addWidget(QLabel("Status:"))
        self.crm_status = QComboBox()
        self.crm_status.addItems(
            ["all", "new", "called", "interested", "not_interested", "callback"])
        self.crm_status.currentTextChanged.connect(self._refresh_crm)
        frow.addWidget(self.crm_status)
        frow.addStretch()
        lay.addLayout(frow)

        self.crm_table = QTableWidget(0, 5)
        self.crm_table.setHorizontalHeaderLabels(
            ["Phone", "Name", "Company", "Status", "Last Called"])
        self.crm_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.crm_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.crm_table.setAlternatingRowColors(True)
        self.crm_table.verticalHeader().setVisible(False)
        lay.addWidget(self.crm_table, stretch=1)
        self._refresh_crm()

    # ══════════════════════════════════════════════════════════════════════════
    #  SETTINGS TAB
    # ══════════════════════════════════════════════════════════════════════════

    def _build_settings_tab(self):
        lay = QVBoxLayout(self.tab_settings)
        lay.setSpacing(6)
        lay.setContentsMargins(10, 8, 10, 8)

        if self.user["role"] != "admin" or self._client_workstation:
            self._build_settings_agent(lay)
            return

        grp_d = QGroupBox("General")
        grp_d.setSizePolicy(
            QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        dlay = QHBoxLayout(grp_d)
        dlay.setContentsMargins(8, 4, 8, 4)
        dlay.setSpacing(8)
        dlay.addWidget(QLabel("Default lines:"))
        self.settings_slots = QSpinBox()
        self.settings_slots.setRange(1, 5)
        self.settings_slots.setValue(self.cfg.get("n_slots", 5))
        dlay.addWidget(self.settings_slots)
        self.chk_ai_audio = QCheckBox("AI audio detection")
        self.chk_ai_audio.setChecked(bool(self.cfg.get("enable_ai_audio", True)))
        dlay.addWidget(self.chk_ai_audio)
        dlay.addWidget(QLabel("Audio device:"))
        self.audio_device_combo = QComboBox()
        try:
            from src.audio_analyzer import AudioAnalyzer
            recommended = AudioAnalyzer.recommend_capture_device()
            if recommended:
                for dev in AudioAnalyzer.list_audio_devices():
                    if str(dev.get("index")) == recommended:
                        self.audio_device_combo.addItem(
                            f"{dev['index']}: {dev['name']} (recommended)",
                            recommended,
                        )
                        break
            self.audio_device_combo.addItem("Auto-detect (Stereo Mix / loopback)", "")
            for dev in AudioAnalyzer.list_audio_devices():
                if dev.get("max_output_channels", 0) or dev.get("max_input_channels", 0):
                    data = str(dev["index"])
                    if self.audio_device_combo.findData(data) >= 0:
                        continue
                    label = f"{dev['index']}: {dev['name']}"
                    self.audio_device_combo.addItem(label, data)
        except Exception:
            self.audio_device_combo.addItem("NO BACKEND", "")
        saved_device = str(self.cfg.get("audio_device", "") or "")
        idx = self.audio_device_combo.findData(saved_device)
        if idx >= 0:
            self.audio_device_combo.setCurrentIndex(idx)
        self.audio_device_combo.setMinimumWidth(200)
        dlay.addWidget(self.audio_device_combo)
        self.chk_live_debug = QCheckBox("Live debug")
        self.chk_live_debug.setChecked(bool(self.cfg.get("live_debug_mode", False)))
        dlay.addWidget(self.chk_live_debug)
        save_btn = _btn("Save settings", "green")
        save_btn.clicked.connect(self._save_settings)
        dlay.addWidget(save_btn)
        dlay.addStretch()
        lay.addWidget(grp_d)

        grp_b = QGroupBox("Voice connection profiles")
        grp_b.setSizePolicy(
            QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        blay = QHBoxLayout(grp_b)
        blay.setContentsMargins(8, 4, 8, 4)
        blay.setSpacing(8)
        hint = QLabel(
            "Sign-in sessions are stored locally. Use Connect account once per line.")
        hint.setWordWrap(True)
        blay.addWidget(hint, stretch=1)
        open_btn = _btn("Open storage folder", "secondary")
        open_btn.clicked.connect(lambda: os.startfile(CHROME_PROFILES_DIR))
        blay.addWidget(open_btn)
        lay.addWidget(grp_b)

        grp_a = QGroupBox("Google Voice accounts")
        alay = QVBoxLayout(grp_a)
        alay.setContentsMargins(8, 6, 8, 6)
        alay.setSpacing(6)
        self.gv_accounts_table = QTableWidget(0, 5)
        self.gv_accounts_table.setHorizontalHeaderLabels(
            ["Priority", "Name", "Email", "Password", "Profile"])
        self.gv_accounts_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.gv_accounts_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.gv_accounts_table.setAlternatingRowColors(True)
        self.gv_accounts_table.verticalHeader().setVisible(False)
        self.gv_accounts_table.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        _configure_data_table(self.gv_accounts_table, min_visible_rows=10)
        alay.addWidget(self.gv_accounts_table, stretch=1)

        acct_buttons = QHBoxLayout()
        acct_buttons.setSpacing(6)
        for txt, fn, nm in [
            ("Add account", self._gv_add_account, "green"),
            ("Connect account", self._gv_setup_selected, "primary"),
            ("Move up", self._gv_move_up, "secondary"),
            ("Move down", self._gv_move_down, "secondary"),
            ("Duplicate", self._gv_duplicate_selected, "secondary"),
            ("Remove", self._gv_remove_selected, "red"),
            ("Refresh", self._refresh_gv_accounts, "secondary"),
        ]:
            b = _btn(txt, nm)
            b.clicked.connect(fn)
            acct_buttons.addWidget(b)
        acct_buttons.addStretch()
        alay.addLayout(acct_buttons)
        lay.addWidget(grp_a, stretch=1)
        self._refresh_gv_accounts()

    def _build_settings_agent(self, lay: QVBoxLayout) -> None:
        grp = QGroupBox("Your account")
        gl = QVBoxLayout(grp)
        gl.setContentsMargins(8, 6, 8, 6)
        gl.addWidget(QLabel(
            "Google Voice lines are configured by your administrator.\n"
            "Use the Dialer and Live Calls tabs to work. "
            "Click Listen on a line to hear the call through your speakers."
        ))
        lay.addWidget(grp)
        lay.addStretch()

    # ══════════════════════════════════════════════════════════════════════════
    #  ADMIN TAB
    # ══════════════════════════════════════════════════════════════════════════

    def _build_admin_tab(self):
        lay = QVBoxLayout(self.tab_admin)
        lay.setSpacing(8)
        lay.setContentsMargins(16, 14, 16, 14)

        top = QHBoxLayout()
        top.addWidget(_label("User Management", bold=True, size=12))
        top.addStretch()
        for txt, fn, nm in [
            ("Add user", self._admin_create, "green"),
            ("Reset password", self._admin_reset_pw, "secondary"),
            ("Set subscription", self._admin_set_subscription, "secondary"),
            ("Activate / deactivate", self._admin_toggle_active, "secondary"),
            ("Delete user", self._admin_delete, "red"),
            ("Refresh", self._admin_refresh, "secondary"),
        ]:
            b = _btn(txt, nm); b.clicked.connect(fn); top.addWidget(b)
        lay.addLayout(top)

        self.admin_table = QTableWidget(0, 9)
        self.admin_table.setHorizontalHeaderLabels(
            ["ID", "Email", "Name", "Role", "Active", "Plan",
             "Expires", "Slots", "Last Login"])
        self.admin_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.admin_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.admin_table.setAlternatingRowColors(True)
        self.admin_table.verticalHeader().setVisible(False)
        lay.addWidget(self.admin_table, stretch=1)
        self._admin_refresh()

        lay.addWidget(_hline())
        grp_client = QGroupBox("Client workstation install")
        gl = QVBoxLayout(grp_client)
        gl.addWidget(QLabel(
            "Use this on YOUR computer to build a folder for the client's PC. "
            "The client will only see agent sign-in — no administrator setup."
        ))
        export_btn = _btn("Export client package…", "primary")
        export_btn.clicked.connect(self._export_client_package)
        gl.addWidget(export_btn)
        lay.addWidget(grp_client)

    def _export_client_package(self):
        from PyQt6.QtWidgets import QInputDialog

        name, ok = QInputDialog.getText(
            self, "Client name", "Client / agent full name:")
        if not ok or not name.strip():
            return
        email, ok = QInputDialog.getText(
            self, "Client email", "Login email for the client:")
        if not ok or not email.strip():
            return
        pw, ok = QInputDialog.getText(
            self, "Client password",
            "Password the client will use to sign in:",
            QLineEdit.EchoMode.Password,
        )
        if not ok or not pw.strip():
            return
        if len(pw) < 8:
            QMessageBox.warning(self, "Password", "Use at least 8 characters.")
            return
        plan, ok = QInputDialog.getItem(
            self, "Subscription", "Client subscription period:",
            ["15 days", "Weekly", "Monthly", "No expiry"], 0, False)
        if not ok:
            return
        days = {"15 days": 15, "Weekly": 7, "Monthly": 30}.get(plan)
        expires_at = ""
        if days:
            expires_at = (
                datetime.now() + timedelta(days=days)
            ).strftime("%Y-%m-%d %H:%M:%S")
        max_slots, ok = QInputDialog.getInt(
            self, "Max live slots", "Allowed simultaneous slots:",
            min(5, max(1, int(self.cfg.get("n_slots", 1)))), 1, 5)
        if not ok:
            return

        out = QFileDialog.getExistingDirectory(
            self, "Save client package to folder (e.g. Desktop)")
        if not out:
            return

        try:
            pkg = export_client_package(
                out, name.strip(), email.strip(), pw, self.cfg,
                copy_voice_profiles=True,
                subscription_plan=plan,
                subscription_expires_at=expires_at,
                max_slots=max_slots,
            )
        except Exception as e:
            QMessageBox.critical(self, "Export failed", str(e))
            return

        QMessageBox.information(
            self, "Client package ready",
            f"Created:\n{pkg}\n\n"
            "1. Install the Auto Dialer app on the client PC\n"
            "2. Copy everything inside that folder into the app folder\n"
            "3. Give the client ONLY their email and password\n\n"
            f"Details are in CLIENT_SETUP.txt",
        )

    # ══════════════════════════════════════════════════════════════════════════
    #  GOOGLE VOICE ACCOUNTS
    # ══════════════════════════════════════════════════════════════════════════

    def _slot_account(self, slot_id: int) -> dict | None:
        if 0 <= slot_id < len(self._gv_accounts):
            return self._gv_accounts[slot_id]
        return None

    def _slot_label(self, slot_id: int) -> str:
        acct = self._slot_account(slot_id)
        if acct:
            return acct.get("name") or acct.get("email") or f"Slot {slot_id + 1}"
        return f"Slot {slot_id + 1}"

    def _selected_gv_account_index(self) -> int:
        if not hasattr(self, "gv_accounts_table"):
            return -1
        row = self.gv_accounts_table.currentRow()
        return row if 0 <= row < len(self._gv_accounts) else -1

    def _refresh_gv_accounts(self):
        self._gv_accounts = load_gv_accounts()
        if not hasattr(self, "gv_accounts_table"):
            return
        theme = self.cfg.get("theme", DEFAULT_THEME)
        self.gv_accounts_table.setRowCount(0)
        for idx, acct in enumerate(self._gv_accounts, start=1):
            row = self.gv_accounts_table.rowCount()
            self.gv_accounts_table.insertRow(row)
            vals = [
                str(idx),
                acct.get("name", ""),
                acct.get("email", ""),
                "Saved" if acct.get("password") else "Manual",
                acct.get("profile", ""),
            ]
            for col, val in enumerate(vals):
                self.gv_accounts_table.setItem(row, col, _table_item(val, theme))
        self._refresh_slot_titles()

    def _refresh_slot_titles(self):
        if not hasattr(self, "_slot_cards"):
            return
        for sid, card in self._slot_cards.items():
            card.set_line_label(self._slot_label(sid))

    def _gv_add_account(self):
        if self.user["role"] != "admin":
            QMessageBox.information(
                self, "Administrator only",
                "Only an administrator can add Google Voice accounts.")
            return
        dlg = AddGVAccountDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        data = dlg.account_data()
        existing = {a.get("profile", "") for a in self._gv_accounts}
        acct = {
            "name": data["name"],
            "email": data["email"],
            "password": data["password"],
            "profile": make_profile_name(data["name"], data["email"], existing),
            "notes": data["notes"],
        }
        self._gv_accounts.append(acct)
        save_gv_accounts(self._gv_accounts)
        self._refresh_gv_accounts()
        self._log(f"Voice account added: {acct['name']}")

        if not self._running:
            n = max(self.spin_slots.value(), len(self._gv_accounts))
            self._init_controllers(n)

        if dlg.auto_login() and acct.get("password"):
            self._start_headless_login(acct)
            QMessageBox.information(
                self, "Signing in",
                f"{acct['name']} is signing in automatically.\n\n"
                "If Google asks for verification, open Settings → Connect account. "
                "Otherwise wait a minute and check Live Calls for Ready status.",
            )
        else:
            self.gv_accounts_table.selectRow(len(self._gv_accounts) - 1)
            self._gv_setup_selected()

    def _gv_move_up(self):
        idx = self._selected_gv_account_index()
        if idx <= 0:
            return
        self._gv_accounts[idx - 1], self._gv_accounts[idx] = (
            self._gv_accounts[idx], self._gv_accounts[idx - 1])
        save_gv_accounts(self._gv_accounts)
        self._refresh_gv_accounts()
        self.gv_accounts_table.selectRow(idx - 1)
        if not self._running:
            self._init_controllers(self.spin_slots.value())

    def _gv_move_down(self):
        idx = self._selected_gv_account_index()
        if idx < 0 or idx >= len(self._gv_accounts) - 1:
            return
        self._gv_accounts[idx + 1], self._gv_accounts[idx] = (
            self._gv_accounts[idx], self._gv_accounts[idx + 1])
        save_gv_accounts(self._gv_accounts)
        self._refresh_gv_accounts()
        self.gv_accounts_table.selectRow(idx + 1)
        if not self._running:
            self._init_controllers(self.spin_slots.value())

    def _gv_duplicate_selected(self):
        idx = self._selected_gv_account_index()
        if idx < 0:
            QMessageBox.warning(self, "Select account", "Select an account first.")
            return
        src = self._gv_accounts[idx]
        existing = {a.get("profile", "") for a in self._gv_accounts}
        copy = dict(src)
        copy["name"] = f"{src.get('name', 'Account')} (copy)"
        copy["profile"] = make_profile_name(copy["name"], src.get("email", ""), existing)
        src_dir = gv_profile_dir(src["profile"])
        dst_dir = gv_profile_dir(copy["profile"])
        if os.path.isdir(src_dir):
            if clone_profile_folder(src["profile"], copy["profile"]):
                self._log(
                    f"Copied signed-in session from {src.get('name')} — no login needed")
            else:
                self._log("Could not copy session — connect the new copy manually")
        self._gv_accounts.insert(idx + 1, copy)
        save_gv_accounts(self._gv_accounts)
        self._refresh_gv_accounts()
        self.gv_accounts_table.selectRow(idx + 1)
        if not self._running:
            self._init_controllers(self.spin_slots.value())
        else:
            self._refresh_slot_login_badges()
        if gv_has_session_marker(dst_dir):
            QMessageBox.information(
                self, "Account duplicated",
                f"{copy['name']} reuses the same Google sign-in — ready to dial.",
            )

    def _gv_remove_selected(self):
        idx = self._selected_gv_account_index()
        if idx < 0:
            QMessageBox.warning(self, "Select Account", "Select an account first.")
            return
        acct = self._gv_accounts[idx]
        if QMessageBox.question(
            self, "Remove Account",
            f"Remove {acct.get('name', 'this account')} from the app?\n\n"
            "The saved browser profile folder is left on disk.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        ) != QMessageBox.StandardButton.Yes:
            return
        self._gv_accounts.pop(idx)
        save_gv_accounts(self._gv_accounts)
        self._refresh_gv_accounts()
        if not self._running:
            self._init_controllers(self.spin_slots.value())

    def _gv_setup_selected(self):
        if self.user["role"] != "admin":
            QMessageBox.information(
                self, "Administrator only",
                "Only an administrator can connect Google Voice accounts.")
            return
        idx = self._selected_gv_account_index()
        if idx < 0:
            QMessageBox.warning(self, "Select account", "Select an account first.")
            return
        acct = self._gv_accounts[idx]
        target_dir = gv_profile_dir(acct["profile"])

        ctrl = next(
            (c for c in self._controllers
             if os.path.abspath(c.profile_dir) == os.path.abspath(target_dir)),
            None
        )
        created_temp = False
        if ctrl is None:
            ctrl = GVController(
                idx,
                target_dir,
                parent=self,
                profile_key=acct["profile"],
                login_email=acct.get("email", ""),
                login_password=acct.get("password", ""),
                runtime_cfg=self._gv_runtime_cfg(),
            )
            ctrl.login_detected.connect(self._on_slot_login)
            ctrl.log_message.connect(self._on_slot_log)
            ctrl.detection_update.connect(self._on_slot_detection)
            created_temp = True
        else:
            ctrl.set_login_credentials(
                acct.get("email", ""),
                acct.get("password", ""),
            )

        acct_label = acct.get("name") or acct.get("email", "Account")

        def _save_password(pw: str) -> None:
            self._gv_accounts[idx]["password"] = pw
            save_gv_accounts(self._gv_accounts)
            acct["password"] = pw
            self._refresh_gv_accounts()
            ctrl.set_login_credentials(acct.get("email", ""), pw)

        dlg = GVSetupDialog(
            ctrl,
            acct_label,
            target_dir,
            login_email=acct.get("email", ""),
            on_password_saved=_save_password,
            main_window=self,
            parent=self,
        )
        dlg.exec()

        if ctrl.is_session_ready() or has_session_marker(target_dir):
            write_session_marker(target_dir)

        if created_temp:
            self._dispose_controller(ctrl)
        else:
            self._hide_browser_after_setup(ctrl.view)

        if not self._running:
            self._init_controllers(self.spin_slots.value())
        else:
            for c in self._controllers:
                if c is None:
                    continue
                if os.path.abspath(c.profile_dir) == os.path.abspath(target_dir):
                    c.mark_logged_in()
                    c.load()

        self._refresh_slot_login_badges()
        self._log(f"Login setup checked for {acct.get('name', acct['email'])}")

        if has_session_marker(target_dir) or ctrl.is_session_ready():
            QMessageBox.information(
                self,
                "Ready to dial",
                f"{acct_label} is connected to Google Voice.\n\n"
                "You can start power dialing — no need to sign in again.",
            )
            if hasattr(self, "tabs"):
                self.tabs.setCurrentWidget(self.tab_live)

    # ══════════════════════════════════════════════════════════════════════════
    #  CONTROLLER INIT
    # ══════════════════════════════════════════════════════════════════════════

    def _init_controllers(self, n: int):
        # Remove old controllers
        for ctrl in self._controllers:
            if ctrl is None:
                continue
            self._watchdog.unregister_slot(ctrl.slot_id)
            self._dispose_controller(ctrl)
        self._controllers.clear()
        QApplication.processEvents()

        for i in range(n):
            acct = self._slot_account(i)
            if acct:
                profile_name = acct["profile"]
                profile_dir = gv_profile_dir(profile_name)
                login_email = acct.get("email", "")
                login_password = acct.get("password", "")
            else:
                profile_name = f"slot_{i}"
                profile_dir = os.path.join(CHROME_PROFILES_DIR, profile_name)
                login_email = ""
                login_password = ""
            ctrl = GVController(i, profile_dir, parent=self,
                                profile_key=profile_name,
                                login_email=login_email,
                                login_password=login_password,
                                runtime_cfg=self._gv_runtime_cfg())
            ctrl.state_changed.connect(self._on_slot_state)
            ctrl.login_detected.connect(self._on_slot_login)
            ctrl.log_message.connect(self._on_slot_log)
            ctrl.heartbeat.connect(self._on_slot_heartbeat)
            ctrl.prepare_for_background_rendering()
            self._controllers.append(ctrl)
            self._watchdog.register_slot(i)
            delay_ms = i * 2500
            QTimer.singleShot(
                delay_ms,
                lambda c=ctrl, pdir=profile_dir: self._boot_controller(c, pdir),
            )
        self._refresh_slot_login_badges()

    def _boot_controller(self, ctrl: GVController, profile_dir: str) -> None:
        """Load Google Voice for one line (staggered so startup does not freeze)."""
        if (
            ctrl.slot_id >= len(self._controllers)
            or self._controllers[ctrl.slot_id] is not ctrl
        ):
            return
        ctrl.load()
        if has_session_marker(profile_dir):
            ctrl.mark_logged_in()
        else:
            QTimer.singleShot(2000, ctrl._check_login)
        self._on_controller_booted()

    def _close_slot_monitor(self, slot_id: int) -> None:
        """Close Listen/debug monitor so it does not hold a disposed WebEngine view."""
        dlg = self._slot_monitors.pop(slot_id, None)
        if dlg is not None:
            try:
                dlg.close()
            except Exception:
                pass

    def _on_slot_heartbeat(self, slot_id: int) -> None:
        self._watchdog.heartbeat(slot_id)

    def _restart_slot(self, slot_id: int, reason: str) -> None:
        """Watchdog: recreate one WebEngine slot without restarting the whole app."""
        import time as _t
        now = _t.time()
        if now - self._slot_restart_cooldown.get(slot_id, 0) < 60:
            return
        if slot_id in self._pending_slot_restarts:
            return
        self._slot_restart_cooldown[slot_id] = now
        log_warning(f"Restarting slot {slot_id}: {reason}")
        self._log(f"[Slot {slot_id}] Recovering line ({reason})…")
        self.statusBar().showMessage(f"Line {slot_id + 1} recovering…")

        phone = self._slot_phone.get(slot_id, "")
        name = self._slot_name.get(slot_id, "")
        attempt = self._slot_retry_attempt.get(slot_id, 0)
        was_running = self._running

        self._update_card(slot_id, "RECOVERING", phone)

        self._recovering_slots.add(slot_id)
        self._close_slot_monitor(slot_id)

        ctrl = self._get_ctrl(slot_id)
        if ctrl:
            try:
                ctrl.hangup()
            except Exception:
                pass
            ctrl.stop_polling()
            self._dispose_controller(ctrl)
            if slot_id < len(self._controllers):
                self._controllers[slot_id] = None

        self._watchdog.unregister_slot(slot_id)

        self._pending_slot_restarts[slot_id] = {
            "phone": phone,
            "name": name,
            "attempt": attempt,
            "was_running": was_running,
        }
        QTimer.singleShot(1200, lambda sid=slot_id: self._finish_slot_restart(sid))

    def _finish_slot_restart(self, slot_id: int) -> None:
        ctx = self._pending_slot_restarts.pop(slot_id, None)
        if ctx is None:
            return

        QApplication.processEvents()

        acct = self._slot_account(slot_id)
        if acct:
            profile_name = acct["profile"]
            profile_dir = gv_profile_dir(profile_name)
            login_email = acct.get("email", "")
            login_password = acct.get("password", "")
        else:
            profile_name = f"slot_{slot_id}"
            profile_dir = os.path.join(CHROME_PROFILES_DIR, profile_name)
            login_email = ""
            login_password = ""

        new_ctrl = GVController(
            slot_id, profile_dir, parent=self,
            profile_key=profile_name,
            login_email=login_email,
            login_password=login_password,
            runtime_cfg=self._gv_runtime_cfg(),
        )
        new_ctrl.state_changed.connect(self._on_slot_state)
        new_ctrl.login_detected.connect(self._on_slot_login)
        new_ctrl.log_message.connect(self._on_slot_log)
        new_ctrl.detection_update.connect(self._on_slot_detection)
        new_ctrl.heartbeat.connect(self._on_slot_heartbeat)
        new_ctrl.prepare_for_background_rendering()

        while len(self._controllers) <= slot_id:
            self._controllers.append(None)
        self._controllers[slot_id] = new_ctrl

        self._watchdog.register_slot(slot_id)
        self._watchdog.reset_call_counter(slot_id)
        new_ctrl.load()
        if has_session_marker(profile_dir):
            new_ctrl.mark_logged_in()

        self._slot_phone.pop(slot_id, None)
        self._slot_start.pop(slot_id, None)
        self._slot_name.pop(slot_id, None)
        self._slot_retry_attempt.pop(slot_id, None)
        self._manual_test_slots.discard(slot_id)
        self._recovering_slots.discard(slot_id)
        self._update_card(slot_id, "IDLE", "")
        self._refresh_slot_login_badges()
        self._update_health_panel()

        phone = ctx.get("phone", "")
        name = ctx.get("name", "")
        attempt = ctx.get("attempt", 0)
        was_running = ctx.get("was_running", False)

        if phone and was_running:
            if self._retry_queue.defer(phone, name, attempt):
                self._log(f"[Slot {slot_id}] Queued retry for {phone}")
            else:
                self._log_final_call(slot_id, "FAILED", phone, contact_name=name)

        self.statusBar().showMessage("Ready")
        if was_running:
            self._schedule_assign(int(self._cooldown_sec() * 500))

    # ══════════════════════════════════════════════════════════════════════════
    #  DIALING LOGIC
    # ══════════════════════════════════════════════════════════════════════════

    def _cooldown_sec(self) -> float:
        return max(1.0, float(self.cfg.get("cooldown", 4.0)))

    def _dial_stagger_sec(self) -> float:
        return max(0.4, float(self.cfg.get("dial_stagger_sec", 0.8)))

    def _slot_is_ready(self, slot_id: int) -> bool:
        return _now() >= self._slot_cooldown_until.get(slot_id, 0)

    def _slot_has_active_phone(self, slot_id: int) -> bool:
        return bool(self._slot_phone.get(slot_id))

    def _controller_available(self, ctrl: GVController | None) -> bool:
        if ctrl is None:
            return False
        sid = ctrl.slot_id
        if sid in self._recovering_slots or sid in self._pending_slot_restarts:
            return False
        if not self._slot_account(sid):
            return False
        if self._slot_has_active_phone(ctrl.slot_id):
            return False
        if ctrl.current_state not in ("IDLE", "ENDED", "ENDED_MANUALLY", "BUSY", "FAILED"):
            return False
        return self._slot_is_ready(ctrl.slot_id)

    def _begin_slot_cooldown(self, slot_id: int, seconds: float | None = None) -> None:
        wait = self._cooldown_sec() if seconds is None else max(0.0, seconds)
        self._slot_cooldown_until[slot_id] = _now() + wait

    def _refresh_dialer_counts(self) -> None:
        if not hasattr(self, "lbl_done"):
            return
        completed = self.db.get_completed_phones()
        done = len(completed)
        total = max(self.progress.maximum() if hasattr(self, "progress") else 0, done)
        active = len(self._slot_phone)
        remaining = max(
            0,
            len(self._contacts) - self._contact_idx
            + self._retry_queue.pending_count()
            + active,
        )
        self.lbl_done.setText(f"Completed: {done}")
        self.lbl_rem.setText(f"Remaining: {remaining}")
        if hasattr(self, "progress"):
            self.progress.setMaximum(total)
            self.progress.setValue(min(done, total))

    def _schedule_assign(self, delay_ms: int | None = None) -> None:
        """Debounced queue pull — avoids hammering all lines at once."""
        if not self._running:
            return
        if delay_ms is None:
            delay_ms = max(600, int(self._cooldown_sec() * 400))
        self._assign_debounce.start(delay_ms)

    def _finish_slot_call(self, slot_id: int) -> None:
        """Line is free: clear UI state, pause, then queue the next number."""
        if not self._running and not self._slot_has_active_phone(slot_id):
            return
        ctrl = self._get_ctrl(slot_id)
        if ctrl:
            ctrl.set_audio_muted(True)
        self._slot_phone.pop(slot_id, None)
        self._slot_start.pop(slot_id, None)
        self._slot_name.pop(slot_id, None)
        self._slot_retry_attempt.pop(slot_id, None)
        self._manual_test_slots.discard(slot_id)
        self._update_card(slot_id, "IDLE", "")
        self._refresh_dialer_counts()
        if not self._running:
            return
        self._begin_slot_cooldown(slot_id)
        self._schedule_assign()

    def _save_campaign_progress(self) -> None:
        self.cfg["campaign_contact_idx"] = int(self._contact_idx)
        path = self.excel_input.text().strip() if hasattr(self, "excel_input") else ""
        self.cfg["campaign_excel_path"] = path.replace("\\", "/")
        _save_cfg(self.cfg)

    def _maybe_offer_campaign_resume(self) -> bool:
        saved_idx = int(self.cfg.get("campaign_contact_idx", 0))
        if saved_idx <= 0 or not self._contacts or saved_idx >= len(self._contacts):
            return False
        saved_path = str(self.cfg.get("campaign_excel_path", "")).replace("\\", "/")
        current_path = self.excel_input.text().strip().replace("\\", "/")
        if saved_path and current_path and saved_path != current_path:
            return False
        ans = QMessageBox.question(
            self,
            "Resume campaign?",
            f"Continue from contact #{saved_idx + 1} of {len(self._contacts)}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        if ans == QMessageBox.StandardButton.Yes:
            self._contact_idx = saved_idx
            self._log(f"Resuming campaign at contact #{saved_idx + 1}")
            return True
        return False

    def _check_stuck_dials(self) -> None:
        if not self._running:
            return
        for sid, started in list(self._slot_start.items()):
            if _now() - started < 35:
                continue
            ctrl = self._get_ctrl(sid)
            if ctrl is None or ctrl.current_state != "DIALING":
                continue
            phone = self._slot_phone.get(sid, "")
            if not phone:
                continue
            self._log(f"[Slot {sid}] Dial stuck 35s+ — auto retrying Call button…")
            ctrl.retry_start_call()

        if self._dial_started_at <= 0:
            return
        elapsed = _now() - self._dial_started_at
        if elapsed < 45:
            return
        active = len(self._slot_phone)
        if active > 0 or self._dial_assign_count > 0:
            return
        busy_states = {
            "DIALING", "RINGING", "ANSWERED_PENDING", "CONNECTED", "VOICEMAIL", "LOADING",
        }
        if any(c and c.current_state in busy_states for c in self._controllers):
            return
        remaining = max(0, len(self._contacts) - self._contact_idx)
        if remaining <= 0 and self._retry_queue.pending_count() <= 0:
            return
        not_ready = []
        for i, ctrl in enumerate(self._controllers):
            if ctrl is None:
                not_ready.append(f"Line {i + 1} (not loaded)")
                continue
            if ctrl.current_state not in ("IDLE", "ENDED", "ENDED_MANUALLY", "READY"):
                not_ready.append(
                    f"Line {i + 1} ({status_label(ctrl.current_state)})")
            elif not self._slot_is_ready(i):
                wait = int(self._slot_cooldown_until.get(i, 0) - _now())
                not_ready.append(f"Line {i + 1} (cooldown {max(wait, 0)}s)")
        hint = (
            "Dialing waiting — no calls assigned yet. "
            "Check Google Voice lines are Ready in Live Calls."
        )
        if not_ready:
            hint += " " + "; ".join(not_ready[:3])
        self.statusBar().showMessage(hint)
        if int(elapsed) % 45 < 11:
            self._log(hint)

    def _retry_slot_dial(self, slot_id: int) -> None:
        ctrl = self._get_ctrl(slot_id)
        if ctrl is None:
            return
        self._log(f"[Slot {slot_id}] Manual retry dial")
        ctrl.retry_start_call()
        if slot_id in self._slot_cards:
            self._slot_cards[slot_id].set_dial_detail("Retrying Call button…")

    def _browse(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "",
            "Excel Files (*.xlsx *.xls);;All Files (*)")
        if path:
            self.excel_input.setText(path)
            self.cfg["excel_path"] = path
            _save_cfg(self.cfg)

    def _load_test_numbers(self):
        """Load built-in owner test list (phones_test.xlsx in project root)."""
        path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "phones_test.xlsx")
        if not os.path.exists(path):
            import subprocess
            script = os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                "scripts", "prepare_test_dial.py")
            try:
                subprocess.run(
                    [sys.executable, script],
                    check=True,
                    cwd=os.path.dirname(os.path.abspath(__file__)),
                )
            except Exception as exc:
                QMessageBox.warning(
                    self, "Test List Missing",
                    f"Could not create test list:\n{exc}")
                return
        self.excel_input.setText(path)
        self.cfg["excel_path"] = path.replace("\\", "/")
        _save_cfg(self.cfg)
        self._load_numbers()

    def _load_numbers(self):
        path = self.excel_input.text().strip()
        if not path or not os.path.exists(path):
            QMessageBox.critical(self, "File Not Found",
                                 f"File not found:\n{path}")
            return
        try:
            df = pd.read_excel(path)
        except Exception as e:
            QMessageBox.critical(self, "Excel Error", str(e))
            return

        df.columns = df.columns.str.strip()
        phone_col = None
        for col in df.columns:
            if col.strip().lower() in ("phone", "mobile", "number",
                                        "tel", "telephone", "cell",
                                        "phone number"):
                phone_col = col; break
        if not phone_col:
            QMessageBox.critical(
                self, "Column Not Found",
                f"No phone column.\nColumns: {list(df.columns)}")
            return

        name_col = next((c for c in df.columns
                         if c.strip().lower() in ("name", "full name",
                                                   "contact name")), None)
        company_col = next((c for c in df.columns
                            if c.strip().lower() in ("company", "business",
                                                      "organization")), None)
        email_col = next((c for c in df.columns
                          if c.strip().lower() in ("email", "e-mail",
                                                    "email address")), None)
        completed = self.db.get_completed_phones()
        valid, invalid = [], 0
        crm_rows: list[dict] = []

        for _, row in df.iterrows():
            d10 = clean_phone(row[phone_col])
            if not d10:
                s = str(row[phone_col]).strip()
                if s.lower() not in ("nan", "none", ""):
                    invalid += 1
                continue
            phone = fmt_e164(d10)
            name  = ""
            if name_col:
                name = str(row[name_col]).strip()
                if name.lower() in ("nan", "none"):
                    name = ""
            company = ""
            if company_col:
                company = str(row[company_col]).strip()
                if company.lower() in ("nan", "none"):
                    company = ""
            email = ""
            if email_col:
                email = str(row[email_col]).strip()
                if email.lower() in ("nan", "none"):
                    email = ""
            crm_rows.append({
                "phone": phone,
                "name": name,
                "company": company,
                "email": email,
            })
            if phone not in completed:
                valid.append((phone, name))

        if not valid:
            QMessageBox.warning(self, "No Numbers",
                                "No valid undialed numbers found.")
            return

        self._contacts    = valid
        self._contact_idx = 0
        done  = len(completed)
        total = len(valid) + done

        self.lbl_total.setText(f"Total: {total}")
        self.lbl_done.setText(f"Completed: {done}")
        self.lbl_rem.setText(f"Remaining: {len(valid)}")
        self.lbl_invalid.setText(f"Invalid: {invalid}")
        self.progress.setMaximum(total)
        self.progress.setValue(done)

        if crm_rows:
            self.db.import_contacts_from_list(crm_rows)
            self._refresh_crm()

        self._log(
            f"Loaded {len(valid)} contacts (completed: {done}, invalid: {invalid})")
        self.btn_start.setEnabled(True)

    def _load_crm_numbers(self, silent: bool = False) -> bool:
        completed = self.db.get_completed_phones()
        rows = sorted(self.db.get_contacts("all"), key=lambda r: int(r.get("id") or 0))
        valid: list[tuple[str, str]] = []
        skip_status = {"not_interested", "do_not_call", "dnc"}
        for row in rows:
            phone = str(row.get("phone") or "").strip()
            if not phone or phone in completed:
                continue
            status = str(row.get("status") or "new").strip().lower()
            if status in skip_status:
                continue
            name = str(row.get("name") or "").strip()
            valid.append((phone, name))

        if not valid:
            if not silent:
                QMessageBox.warning(
                    self,
                    "No CRM Contacts",
                    "No dialable CRM contacts found. Import or load an Excel list first.",
                )
            return False

        self._contacts = valid
        self._contact_idx = 0
        done = len(completed)
        total = len(valid) + done
        self.lbl_total.setText(f"Total: {total}")
        self.lbl_done.setText(f"Completed: {done}")
        self.lbl_rem.setText(f"Remaining: {len(valid)}")
        self.lbl_invalid.setText("Invalid: 0")
        self.progress.setMaximum(total)
        self.progress.setValue(done)
        self.btn_start.setEnabled(True)
        self._log(f"Loaded {len(valid)} CRM contacts (completed: {done})")
        return True

    def _start_dialing(self):
        if not self._contacts:
            if not self._load_crm_numbers(silent=True):
                QMessageBox.warning(
                    self,
                    "No Contacts",
                    "Load an Excel list or import contacts into CRM first.",
                )
                return

        ok, msg = self._dialing_login_ok()
        if not ok:
            self._log(f"Google Voice line readiness failed: {msg}")
            QMessageBox.warning(self, "Google Voice Not Ready", msg)
            return
        self._log("Google Voice line readiness confirmed for selected lines")

        n = self.spin_slots.value()
        self.cfg.update({
            "n_slots": n,
            "call_timeout": self.spin_timeout.value(),
            "cooldown":     self.spin_cooldown.value(),
            "voicemail_hangup_sec": self.spin_vm_hangup.value(),
            "predictive_mode": self.chk_predictive.isChecked() if hasattr(self, "chk_predictive") else False,
        })
        _save_cfg(self.cfg)
        self._maybe_start_websocket()
        cd = self._cooldown_sec()
        self._dial_timer.setInterval(max(2500, int(cd * 650)))

        # Re-init controllers if slot count changed
        if len(self._controllers) != n:
            self._init_controllers(n)
            self._rebuild_slot_cards(n)

        self._campaign_generation += 1
        self._running      = True
        self._dial_assign_count = 0
        self._dial_started_at = _now()
        resume = self._maybe_offer_campaign_resume()
        if not resume:
            self._contact_idx = 0
        self._retry_queue.clear()
        self._slot_cooldown_until.clear()
        stagger = self._dial_stagger_sec()
        for i in range(n):
            self._slot_cooldown_until[i] = _now() + i * stagger
        for ctrl in self._controllers:
            if ctrl:
                ctrl.set_audio_muted(True)
        self._configure_watchdog()
        self._sync_dial_buttons(True)
        self.statusBar().showMessage("Dialing in progress…")
        self._log(
            f"Dialing started — {n} line(s), {cd:.1f}s pause between calls per line")

        self._dial_timer.start()
        self._elapsed_timer.start()
        self._stuck_dial_timer.start()
        self._schedule_assign(int(stagger * 1000))

    def _stop_dialing(self):
        self._campaign_generation += 1
        self._running = False
        self._dial_timer.stop()
        self._assign_debounce.stop()
        self._elapsed_timer.stop()
        self._stuck_dial_timer.stop()
        self._save_campaign_progress()
        pending = self._retry_queue.pending_count()
        self._retry_queue.clear()
        for ctrl in self._controllers:
            if ctrl is None:
                continue
            sid = ctrl.slot_id
            self._slot_call_token[sid] = self._slot_call_token.get(sid, 0) + 1
            ctrl.stop_polling()
            try:
                ctrl.hangup()
            except Exception:
                pass
            self._slot_phone.pop(sid, None)
            self._slot_start.pop(sid, None)
            self._slot_name.pop(sid, None)
            self._slot_retry_attempt.pop(sid, None)
            self._manual_test_slots.discard(sid)
            self._slot_cooldown_until.pop(sid, None)
            self._watchdog.record_state(sid, "IDLE")
            self._update_card(sid, "IDLE", "")
        self._refresh_dialer_counts()
        self._sync_dial_buttons(False)
        self.statusBar().showMessage("Dialing stopped")
        self._log("Dialing stopped")
        log_info(
            f"Campaign stopped — retries still queued: {pending}, "
            f"log: {log_path()}")

    def _emergency_stop(self):
        self._log("EMERGENCY STOP requested - stopping all live dialing now")
        log_warning("Emergency stop requested from UI")
        self._stop_dialing()

    def _assign_pending_calls(self):
        """Assign retries first, then new contacts, to idle lines past cooldown."""
        if not self._running:
            return

        assigned = 0
        max_assign = 999
        if bool(self.cfg.get("predictive_mode", False)):
            stats = self.db.recent_pacing_stats()
            agents = max(1, available_agent_count())
            in_progress = sum(
                1 for c in self._controllers
                if c and c.current_state in (
                    "DIALING", "RINGING", "ANSWERED_PENDING", "CONNECTED", "VOICEMAIL")
            )
            metrics = PacingMetrics(
                agents_available=agents,
                connect_rate=float(stats.get("connect_rate", 0.15)),
                abandon_rate=float(stats.get("abandon_rate", 0.0)),
                calls_in_progress=in_progress,
            )
            max_assign = self._pacing_engine.smoothed_dials_needed(
                metrics, samples=self._pacing_dial_history[-4:],
            )
            self._pacing_dial_history.append(max_assign)
            self._update_pacing_panel(stats, max_assign)
            if hasattr(self, "_ws_thread") and self._ws_thread is not None:
                self._ws_thread.broadcast("pacing_update", {
                    "connect_rate": metrics.connect_rate,
                    "abandon_rate": metrics.abandon_rate,
                    "dials_needed": max_assign,
                })

        ready_retries = self._retry_queue.pop_ready()
        deferred_retries: list[tuple[str, str, int]] = []

        for phone, name, attempt in ready_retries:
            if assigned >= max_assign:
                deferred_retries.append((phone, name, attempt))
                continue
            ctrl = self._idle_controller()
            if ctrl is None:
                deferred_retries.append((phone, name, attempt))
                continue
            self._dial_on_slot(ctrl, phone, name, attempt)
            assigned += 1
            self._dial_assign_count += 1

        for phone, name, attempt in deferred_retries:
            self._retry_queue.requeue(phone, name, attempt, 2.0)

        for ctrl in self._controllers:
            if assigned >= max_assign:
                break
            if not self._controller_available(ctrl):
                continue
            if self._contact_idx >= len(self._contacts):
                break
            phone, name = self._contacts[self._contact_idx]
            self._contact_idx += 1
            if self._contact_idx % 3 == 0:
                self._save_campaign_progress()
            self._dial_on_slot(ctrl, phone, name, 0)
            assigned += 1
            self._dial_assign_count += 1

        if assigned == 0 and self._running:
            next_ready = min(
                (t for t in self._slot_cooldown_until.values() if t > _now()),
                default=0,
            )
            if next_ready > _now():
                self._schedule_assign(int((next_ready - _now()) * 1000) + 200)

        if (self._contact_idx >= len(self._contacts)
                and self._retry_queue.pending_count() == 0):
            all_idle = all(
                c and c.current_state in (
                    "IDLE", "ENDED", "ENDED_MANUALLY", "VOICEMAIL", "NO_ANSWER", "BUSY", "FAILED")
                for c in self._controllers)
            if all_idle:
                self._dial_timer.stop()
                self._elapsed_timer.stop()
                self._on_all_done()

    def _update_pacing_panel(self, stats: dict, dials_needed: int) -> None:
        if not hasattr(self, "lbl_connect_rate"):
            return
        cr = float(stats.get("connect_rate", 0.0)) * 100
        ar = float(stats.get("abandon_rate", 0.0)) * 100
        self.lbl_connect_rate.setText(f"Connect: {cr:.0f}%")
        self.lbl_abandon_rate.setText(f"Abandon: {ar:.1f}%")
        self.lbl_dials_interval.setText(f"Dials/interval: {dials_needed}")

    def _maybe_start_websocket(self) -> None:
        if not bool(self.cfg.get("websocket_enabled", False)):
            return
        if self._ws_thread is not None:
            return
        try:
            from src.ui.websocket_manager import WebSocketServerThread
            port = int(self.cfg.get("websocket_port", 8765))
            self._ws_thread = WebSocketServerThread(port=port, parent=self)
            self._ws_thread.start()
        except Exception as exc:
            log_warning(f"WebSocket dashboard unavailable: {exc}")

    def _broadcast_amd_event(self, slot_id: int, debug: dict) -> None:
        if self._ws_thread is None:
            return
        self._ws_thread.broadcast("amd_decision", {"slot_id": slot_id, **debug})

    def _idle_controller(self) -> GVController | None:
        for ctrl in self._controllers:
            if self._controller_available(ctrl):
                return ctrl
        return None

    def _dial_on_slot(self, ctrl: GVController, phone: str, name: str,
                      retry_attempt: int) -> None:
        sid = ctrl.slot_id
        self._slot_cooldown_until.pop(sid, None)
        self._slot_phone[sid] = phone
        self._slot_name[sid] = name
        self._slot_retry_attempt[sid] = retry_attempt
        self._slot_start[sid] = _now()
        token = self._slot_call_token.get(sid, 0) + 1
        self._slot_call_token[sid] = token
        self._slot_final_logged.discard(sid)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self._slot_call_times[sid] = {"dialed_at": now}
        self._update_card(sid, "DIALING", phone)
        if retry_attempt > 0:
            self._log(
                f"[Slot {sid}] Retry {retry_attempt} — dialing {phone}…")
            log_info(f"Slot {sid} retry #{retry_attempt} for {phone}")
        else:
            self._log(f"[Slot {sid}] Dialing {phone}…")
        try:
            self._log(f"[Slot {sid}] Dial attempt started: {phone}")
            ctrl.dial(phone)
        except Exception:
            tb = traceback.format_exc()
            self._log(f"[Slot {sid}] FAILED: exception while starting dial\n{tb}")
            log_error(f"Slot {sid} dial exception for {phone}\n{tb}")
            self._handle_slot_failure(sid, phone)
            return
        timeout_ms = int(self.cfg.get("call_timeout", 60) * 1000)
        started_at = self._slot_start[sid]
        campaign_generation = self._campaign_generation
        QTimer.singleShot(
            timeout_ms,
            lambda sid=sid, p=phone, st=started_at, tok=token, gen=campaign_generation:
            self._timeout_call(sid, p, st, tok, gen),
        )

    def _release_slot(self, slot_id: int):
        self._next_call(slot_id)

    def _cut_call(self, slot_id: int, advance: bool = True):
        """Hang up the current backend Google Voice call from our UI."""
        ctrl = self._get_ctrl(slot_id)
        phone = self._slot_phone.get(slot_id, "")
        state = ctrl.current_state if ctrl else "IDLE"
        started = self._slot_start.get(slot_id)
        duration_s = max(0.0, _now() - started) if started else 0.0
        if ctrl:
            ctrl.set_audio_muted(True)
            ctrl.hangup(manual=True)
            self._log(f"[Slot {slot_id}] Cut call")
        if phone:
            status = "ENDED_MANUALLY"
            self._log_final_call(
                slot_id, status, phone,
                duration_s=duration_s,
            )
            self._watchdog.record_call_completed(slot_id)
        self._slot_phone.pop(slot_id, None)
        self._slot_start.pop(slot_id, None)
        self._slot_name.pop(slot_id, None)
        self._slot_retry_attempt.pop(slot_id, None)
        self._manual_test_slots.discard(slot_id)
        self._update_card(slot_id, "IDLE", "")
        if advance and self._running:
            self._begin_slot_cooldown(slot_id)
            self._schedule_assign()

    def _next_call(self, slot_id: int):
        """Cut the current call and assign the next number after a short pause."""
        self._cut_call(slot_id, advance=False)
        self._log(f"[Slot {slot_id}] Moving to next call")
        if self._running:
            pause = max(1.0, self._cooldown_sec() * 0.5)
            self._begin_slot_cooldown(slot_id, pause)
            self._schedule_assign(int(pause * 1000))

    def _voicemail_hangup_and_next(
        self,
        slot_id: int,
        campaign_generation: int | None = None,
        token: int | None = None,
    ) -> None:
        """Hang up voicemail and advance the power dialer queue."""
        if campaign_generation is not None and campaign_generation != self._campaign_generation:
            return
        if token is not None and self._slot_call_token.get(slot_id) != token:
            return
        ctrl = self._get_ctrl(slot_id)
        if ctrl and ctrl.current_state == "VOICEMAIL":
            ctrl.set_audio_muted(True)
            ctrl.hangup()
        self._slot_phone.pop(slot_id, None)
        self._slot_start.pop(slot_id, None)
        self._slot_name.pop(slot_id, None)
        self._slot_retry_attempt.pop(slot_id, None)
        self._manual_test_slots.discard(slot_id)
        self._update_card(slot_id, "IDLE", "")
        self._log(f"[Slot {slot_id}] Voicemail handled — next number")
        if self._running:
            self._begin_slot_cooldown(slot_id)
            self._schedule_assign()

    def _timeout_call(self, slot_id: int, phone: str, started_at: float,
                      token: int, campaign_generation: int):
        """Auto-cut an unanswered call once the configured timeout expires."""
        if not self._running:
            return
        if campaign_generation != self._campaign_generation:
            return
        if self._slot_call_token.get(slot_id) != token:
            return
        ctrl = self._get_ctrl(slot_id)
        if not ctrl:
            return
        if self._slot_phone.get(slot_id) != phone:
            return
        if self._slot_start.get(slot_id) != started_at:
            return
        if ui_display_state(self._slot_ui_state.get(slot_id, ctrl.current_state)) == "CONNECTED":
            return
        if ctrl.current_state in ("DIALING", "RINGING", "IDLE"):
            self._log(f"[Slot {slot_id}] Timeout reached — retry or skip")
            self._handle_slot_failure(slot_id, phone)

    def _on_all_done(self):
        self._running = False
        self._stuck_dial_timer.stop()
        self.cfg["campaign_contact_idx"] = 0
        _save_cfg(self.cfg)
        self._sync_dial_buttons(False)
        self.statusBar().showMessage("Campaign complete")
        self._log("All contacts in this list have been dialed.")
        QMessageBox.information(self, "Done", "All contacts have been dialed!")

    # ── Slot state handling ───────────────────────────────────────────────────

    def _log_final_call(self, slot_id: int, status: str, phone: str,
                        *, duration_s: float = 0.0, contact_name: str = "") -> bool:
        if not phone or slot_id in self._slot_final_logged:
            return False
        times = self._slot_call_times.get(slot_id, {})
        debug = self._slot_detection_debug.get(slot_id, {})
        history = json.dumps({
            "dom_state": debug.get("dom_state"),
            "audio_state": debug.get("audio_state"),
            "fused_state": debug.get("fused_state"),
            "elapsed": debug.get("elapsed"),
        })
        self.db.log_call(
            self.user["id"],
            phone,
            status,
            contact_name=contact_name or self._slot_name.get(slot_id, ""),
            duration_s=duration_s,
            slot_id=slot_id,
            detection_reason=str(debug.get("reason") or debug.get("audio_reason") or ""),
            confidence=float(debug.get("confidence") or 0.0),
            state_history=history,
            final_outcome=status,
            dialed_at=times.get("dialed_at", ""),
            ringing_at=times.get("ringing_at", ""),
            connected_at=times.get("connected_at", ""),
            detection_time_ms=int(debug.get("detection_time_ms") or 0),
        )
        self._slot_call_times.pop(slot_id, None)
        self._slot_final_logged.add(slot_id)
        self._refresh_logs()
        self._refresh_dialer_counts()
        return True

    def _handle_slot_failure(self, slot_id: int, phone: str) -> None:
        """Retry with backoff or log FAILED after max attempts."""
        name = self._slot_name.get(slot_id, "")
        attempt = self._slot_retry_attempt.pop(slot_id, 0)
        ctrl = self._get_ctrl(slot_id)
        if ctrl:
            try:
                ctrl.set_audio_muted(True)
                if ctrl.current_state in ("RINGING", "ANSWERED_PENDING", "CONNECTED", "VOICEMAIL"):
                    ctrl.hangup()
            except Exception:
                pass
        if phone:
            if slot_id in self._manual_test_slots:
                self._log_final_call(slot_id, "FAILED", phone, contact_name=name)
            elif self._retry_queue.defer(phone, name, attempt):
                log_info(f"Queued retry for {phone} (after attempt {attempt + 1})")
                self._log(f"[Slot {slot_id}] Will retry {phone} shortly")
            else:
                self._log_final_call(slot_id, "FAILED", phone, contact_name=name)
                log_warning(f"Max retries exhausted for {phone}")
        self._slot_phone.pop(slot_id, None)
        self._slot_start.pop(slot_id, None)
        self._slot_name.pop(slot_id, None)
        self._manual_test_slots.discard(slot_id)
        self._update_card(slot_id, "IDLE", "")
        self._watchdog.record_call_completed(slot_id)
        if self._running:
            self._begin_slot_cooldown(slot_id)
            self._schedule_assign()

    def _on_slot_state(self, slot_id: int, state: str):
        display_state = ui_display_state(state)
        if not self._running and not self._slot_has_active_phone(slot_id):
            if state == "IDLE":
                self._watchdog.record_state(slot_id, "IDLE")
                self._update_card(slot_id, "IDLE", "")
            return

        if state == "IDLE" and self._slot_has_active_phone(slot_id):
            ctrl = self._get_ctrl(slot_id)
            last = ctrl.current_state if ctrl else "DIALING"
            if last == "IDLE":
                last = "DIALING"
            self._update_card(slot_id, last, self._slot_phone.get(slot_id, ""))
            return

        self._watchdog.record_state(slot_id, state)
        phone = self._slot_phone.get(slot_id, "")
        disp  = fmt_display(phone[2:]) if phone.startswith("+1") and len(phone) == 12 \
            else phone
        if display_state == "RINGING":
            self._mark_call_time(slot_id, "ringing_at")
            self._log(f"[Slot {slot_id}] Ringing: {disp}")
            if state == "ANSWERED_PENDING":
                self.statusBar().showMessage(
                    f"Classifying answer — Line {slot_id + 1}: {disp}")
        elif display_state == "CONNECTED":
            self._mark_call_time(slot_id, "connected_at")
            self._log(
                f"[Slot {slot_id}] Call picked up - talk now: {disp} ({state})"
            )
            self.statusBar().showMessage(
                f"Call picked up - talk now - Line {slot_id + 1}: {disp}")
        self._update_card(slot_id, state, phone)
        self._log(f"[Slot {slot_id}] → {state}  {disp}")

        if display_state == "CONNECTED":
            ui_cfg = self.cfg.get("ui") if isinstance(self.cfg.get("ui"), dict) else {}
            panel_delay_ms = int(float(ui_cfg.get("agent_panel_delay_seconds", 0)) * 1000)

            def _show_connected(sid=slot_id, d=disp, delay=panel_delay_ms):
                self._focus_answered_slot(sid)
                self.statusBar().showMessage(
                    f"Call picked up - talk now - Line {sid + 1}: {d}"
                )
                auto_listen = bool(ui_cfg.get("auto_open_panel_on_human", False))
                if auto_listen and effective_enable_ai_audio(self.cfg):
                    self._open_slot_monitor(sid)

            if panel_delay_ms > 0:
                QTimer.singleShot(panel_delay_ms, _show_connected)
            else:
                _show_connected()
            wait_sec = float(self.cfg.get("target_agent_wait_seconds", 3))
            token = self._slot_call_token.get(slot_id, 0)
            gen = self._campaign_generation
            QTimer.singleShot(
                int(wait_sec * 1000),
                lambda sid=slot_id, tok=token, g=gen:
                self._check_abandoned_call(sid, tok, g),
            )
        elif state == "VOICEMAIL":
            vm_sec = int(self.cfg.get("voicemail_hangup_sec", 3))
            self._log(
                f"[Slot {slot_id}] Voicemail detected: {disp} - auto-hangup in {vm_sec}s, then next number"
            )
            self._log_final_call(slot_id, "VOICEMAIL", phone)
            self._watchdog.record_call_completed(slot_id)
            campaign_generation = self._campaign_generation
            token = self._slot_call_token.get(slot_id, 0)
            QTimer.singleShot(
                vm_sec * 1000,
                lambda sid=slot_id, gen=campaign_generation, tok=token:
                self._voicemail_hangup_and_next(sid, gen, tok),
            )
        elif display_state == "FAILED":
            self._log(f"[Slot {slot_id}] failed: {disp}")
            self._handle_slot_failure(slot_id, phone)
        elif display_state == "ENDED" and state != "ENDED_MANUALLY":
            if phone:
                started = self._slot_start.get(slot_id)
                duration_s = max(0.0, _now() - started) if started else 0.0
                self._log_final_call(
                    slot_id, "ENDED", phone,
                    contact_name=self._slot_name.get(slot_id, ""),
                    duration_s=duration_s,
                )
                self._watchdog.record_call_completed(slot_id)
                self._finish_slot_call(slot_id)
        elif display_state == "NO_ANSWER":
            if phone:
                self._log(f"[Slot {slot_id}] No answer / hung up before pickup: {disp}")
                self._log_final_call(
                    slot_id, "NO_ANSWER", phone,
                    contact_name=self._slot_name.get(slot_id, ""),
                )
                self._watchdog.record_call_completed(slot_id)
                self._finish_slot_call(slot_id)
        elif display_state == "BUSY":
            if phone:
                self._log_final_call(
                    slot_id, "BUSY", phone,
                    contact_name=self._slot_name.get(slot_id, ""),
                )
                self._watchdog.record_call_completed(slot_id)
                self._finish_slot_call(slot_id)
        elif state == "ENDED_MANUALLY":
            if phone:
                started = self._slot_start.get(slot_id)
                duration_s = max(0.0, _now() - started) if started else 0.0
                self._log_final_call(
                    slot_id, "ENDED_MANUALLY", phone,
                    contact_name=self._slot_name.get(slot_id, ""),
                    duration_s=duration_s,
                )
                self._watchdog.record_call_completed(slot_id)
                self._finish_slot_call(slot_id)
        elif display_state == "IDLE":
            pass

    def _on_slot_login(self, slot_id: int):
        ctrl = self._get_ctrl(slot_id)
        if ctrl:
            ctrl.mark_logged_in()
        self._log(f"[Slot {slot_id}] Google Voice line readiness: ready")
        self._update_card(slot_id, "IDLE", "")
        self._refresh_slot_login_badges()

    def _on_slot_log(self, slot_id: int, msg: str):
        self._log(f"[Slot {slot_id}] {msg}")
        if "Traceback (most recent call last)" in msg:
            log_error(f"Slot {slot_id} traceback from UI log\n{msg}")
        if "Dial UI status:" in msg and slot_id in getattr(self, "_slot_cards", {}):
            detail = msg.split("Dial UI status:", 1)[-1].strip()
            self._slot_cards[slot_id].set_dial_detail(detail[:80])
        if "sign-in required" in msg.lower() or "google account detected" in msg.lower():
            self._refresh_slot_login_badges()

    def _check_abandoned_call(self, slot_id: int, token: int, generation: int) -> None:
        if generation != self._campaign_generation:
            return
        if token != self._slot_call_token.get(slot_id, 0):
            return
        if slot_id in self._slot_agent_handled:
            return
        ctrl = self._get_ctrl(slot_id)
        if not ctrl or ctrl.current_state != "CONNECTED":
            return
        phone = self._slot_phone.get(slot_id, "")
        self._pacing_engine.record_outcome(connected=True, abandoned=True)
        self.db.log_call(
            self.user["id"],
            phone,
            "ABANDONED",
            contact_name=self._slot_name.get(slot_id, ""),
            slot_id=slot_id,
            detection_reason="human detected but agent not connected in time",
        )
        self._log(f"[Slot {slot_id}] Abandoned — no agent within wait window")

    def _on_slot_detection(self, slot_id: int, debug: dict) -> None:
        self._slot_detection_debug[slot_id] = debug
        if hasattr(self, "_slot_cards") and slot_id in self._slot_cards:
            self._slot_cards[slot_id].update_detection(debug)
        fused_state = str(debug.get("fused_state") or debug.get("ui_state") or "").upper()
        if ui_display_state(fused_state) == "CONNECTED" and self._slot_has_active_phone(slot_id):
            self._update_card(slot_id, fused_state, self._slot_phone.get(slot_id, ""))
        if fused_state in ("HUMAN", "VOICEMAIL", "CONNECTED", "CONNECTED_AUDIO_EVIDENCE", "ANSWERED_PENDING"):
            self._broadcast_amd_event(slot_id, debug)
            if fused_state in {"HUMAN", "CONNECTED", "CONNECTED_AUDIO_EVIDENCE"}:
                self._pacing_engine.record_outcome(connected=True, abandoned=False)

    def _update_card(self, slot_id: int, state: str, phone: str):
        if hasattr(self, "_slot_cards") and slot_id in self._slot_cards:
            display_state = ui_display_state(state)
            current_state = self._slot_ui_state.get(slot_id, "IDLE")
            if not ui_state_allows_transition(current_state, display_state):
                self._log(
                    f"[Slot {slot_id}] UI status kept at {current_state}; ignored stale {state}"
                )
                return
            elapsed = ""
            if slot_id in self._slot_start and display_state not in ("IDLE", "ENDED"):
                s = int(_now() - self._slot_start[slot_id])
                elapsed = f"{s//60:02d}:{s%60:02d}"
            disp = fmt_display(phone[2:]) if phone.startswith("+1") and len(phone)==12 \
                else phone
            self._slot_cards[slot_id].update_state(state, disp, elapsed)
            if self._slot_ui_state.get(slot_id) != display_state:
                self._log(
                    f"[Slot {slot_id}] UI status {self._slot_ui_state.get(slot_id, 'IDLE')} -> {display_state} (backend {state})"
                )
            self._slot_ui_state[slot_id] = display_state
            if display_state == "IDLE":
                self._slot_cards[slot_id].set_dial_detail("")

    def _tick_elapsed(self):
        for ctrl in self._controllers:
            if ctrl is None:
                continue
            sid = ctrl.slot_id
            if ctrl.current_state not in ("IDLE", "ENDED", "ENDED_MANUALLY", "BUSY", "FAILED"):
                state = self._slot_ui_state.get(sid) or ctrl.current_state
                self._update_card(sid, state,
                                  self._slot_phone.get(sid, ""))

    def _get_ctrl(self, slot_id: int) -> GVController | None:
        if 0 <= slot_id < len(self._controllers):
            return self._controllers[slot_id]
        return None

    # ── Logs ─────────────────────────────────────────────────────────────────

    def _refresh_logs(self):
        uid = None if self.user["role"] == "admin" else self.user["id"]
        self._all_logs = self.db.get_call_records(user_id=uid)
        self._apply_log_filter()

    def _apply_log_filter(self):
        q  = self.log_search.text().strip().lower() if hasattr(self, "log_search") else ""
        sf = self.log_status_combo.currentText() if hasattr(self, "log_status_combo") else "All Statuses"
        filtered = []
        for r in self._all_logs:
            st = r.get("status", "")
            if sf != "All Statuses" and st != sf:
                continue
            if q and q not in r.get("phone", "").lower() \
               and q not in st.lower() \
               and q not in r.get("timestamp", "").lower():
                continue
            filtered.append(r)

        ended = vm = fail = 0
        for r in self._all_logs:
            st = r.get("status", "")
            if st == "ENDED":       ended += 1
            elif st == "VOICEMAIL": vm    += 1
            elif st == "FAILED":    fail  += 1
            elif st == "BUSY":      fail  += 1
        self.log_total.setText(f"Total: {len(self._all_logs)}")
        self.log_ended.setText(f"Ended: {ended}")
        self.log_vm.setText(f"Voicemail: {vm}")
        self.log_fail.setText(f"Failed: {fail}")

        self.log_table.setRowCount(0)
        theme = self.cfg.get("theme", DEFAULT_THEME)
        default_fg = QColor(table_item_color(theme))
        for r in reversed(filtered):
            row = self.log_table.rowCount()
            self.log_table.insertRow(row)
            st  = r.get("status", "")
            dur = r.get("duration_s", 0) or 0
            vals = [
                r.get("timestamp", ""),
                r.get("phone", ""),
                st,
                f"{dur:.0f}s",
                f"L{(r.get('slot_id', 0) or 0) + 1}",
                r.get("dialed_at", "") or "—",
                r.get("connected_at", "") or "—",
            ]
            status_fg = QColor(log_status_color(st, theme))
            for col, val in enumerate(vals):
                item = _table_item(val, theme)
                if col == 2 and st:
                    item.setForeground(status_fg)
                else:
                    item.setForeground(default_fg)
                self.log_table.setItem(row, col, item)

    def _export_logs(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Logs",
            f"FTSolutions_CallLog_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            "Excel (*.xlsx);;CSV (*.csv)"
        )
        if not path or not self._all_logs:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Call History"
            headers = ["Time", "Phone", "Status", "Duration (s)", "Slot"]
            ws.append(headers)
            hdr_fill = PatternFill("solid", fgColor="1A7F37")
            hdr_font = Font(bold=True, color="FFFFFF")
            for c in range(1, len(headers)+1):
                ws.cell(1, c).fill = hdr_fill
                ws.cell(1, c).font = hdr_font
            fill_map = {
                "ENDED":     PatternFill("solid", fgColor="0A2010"),
                "VOICEMAIL": PatternFill("solid", fgColor="1A0F00"),
                "NO_ANSWER": PatternFill("solid", fgColor="111820"),
                "FAILED":    PatternFill("solid", fgColor="1A0000"),
            }
            for r in self._all_logs:
                ws.append([r.get("timestamp",""), r.get("phone",""),
                           r.get("status",""), r.get("duration_s",0),
                           r.get("slot_id",0)])
                st = r.get("status","")
                if st in fill_map:
                    for c in range(1, len(headers)+1):
                        ws.cell(ws.max_row, c).fill = fill_map[st]
            wb.save(path)
            QMessageBox.information(self, "Exported", f"Saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    def _clear_logs(self):
        if QMessageBox.question(
                self, "Clear Logs", "Delete ALL call logs?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        ) == QMessageBox.StandardButton.Yes:
            try:
                with self.db._conn() as c:
                    c.execute("DELETE FROM call_records")
                from src.paths import CALL_LOG_CSV
                if os.path.exists(CALL_LOG_CSV):
                    os.remove(CALL_LOG_CSV)
            except Exception:
                pass
            self._all_logs = []
            self._apply_log_filter()
            self._log("🗑 Logs cleared")

    # ── CRM ───────────────────────────────────────────────────────────────────

    def _refresh_crm(self):
        if not hasattr(self, "crm_table"):
            return
        sf = self.crm_status.currentText() if hasattr(self, "crm_status") else "all"
        rows = self.db.get_contacts(sf)
        self.crm_table.setRowCount(0)
        for r in rows:
            row = self.crm_table.rowCount()
            self.crm_table.insertRow(row)
            for col, key in enumerate(("phone","name","company","status","last_called")):
                val = str(r.get(key,"") or "—")
                item = QTableWidgetItem(val)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.crm_table.setItem(row, col, item)

    def _import_contacts(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Contacts", "",
            "Excel Files (*.xlsx *.xls);;All Files (*)")
        if not path:
            return
        try:
            df = pd.read_excel(path)
            df.columns = df.columns.str.strip().str.lower()
            rows = []
            for _, r in df.iterrows():
                for col in ("phone","mobile","number","tel"):
                    if col in df.columns:
                        d10 = clean_phone(r[col])
                        if d10:
                            rows.append({
                                "phone":   fmt_e164(d10),
                                "name":    str(r.get("name","")).strip(),
                                "company": str(r.get("company","")).strip(),
                                "email":   str(r.get("email","")).strip(),
                            })
                            break
            added, skipped = self.db.import_contacts_from_list(rows)
            QMessageBox.information(
                self, "Import Done",
                f"Added: {added}  |  Skipped: {skipped}")
            self._refresh_crm()
        except Exception as e:
            QMessageBox.critical(self, "Import Error", str(e))

    def _add_contact(self):
        from PyQt6.QtWidgets import QInputDialog
        phone, ok = QInputDialog.getText(self, "Add Contact", "Phone Number:")
        if not ok or not phone:
            return
        d10 = clean_phone(phone)
        if not d10:
            QMessageBox.warning(self, "Invalid", "Not a valid US phone number.")
            return
        name, _ = QInputDialog.getText(self, "Add Contact", "Name (optional):")
        self.db.upsert_contact(fmt_e164(d10), name=name.strip())
        self._refresh_crm()

    def _delete_contact(self):
        row = self.crm_table.currentRow()
        if row < 0:
            return
        phone = self.crm_table.item(row, 0).text()
        if QMessageBox.question(
                self, "Delete", f"Delete {phone}?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        ) == QMessageBox.StandardButton.Yes:
            self.db.delete_contact(phone)
            self._refresh_crm()

    # ── Admin ─────────────────────────────────────────────────────────────────

    def _admin_refresh(self):
        if not hasattr(self, "admin_table"):
            return
        self.admin_table.setRowCount(0)
        for u in self.db.get_all_users():
            row = self.admin_table.rowCount()
            self.admin_table.insertRow(row)
            expired = self.db.user_subscription_expired(u)
            active_label = "Expired" if expired else (
                "Active" if u["is_active"] else "Inactive")
            vals = [str(u["id"]), u["email"], u["name"], u["role"],
                    active_label,
                    u.get("subscription_plan", "manual") or "manual",
                    u.get("subscription_expires_at", "") or "No expiry",
                    str(u.get("max_slots", 1) or 1),
                    u.get("last_login","—") or "—"]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if u["role"] == "admin":
                    item.setForeground(QColor("#ffd166"))
                elif expired:
                    item.setForeground(QColor("#ff4444"))
                elif not u["is_active"]:
                    item.setForeground(QColor("#8b949e"))
                self.admin_table.setItem(row, col, item)

    def _admin_create(self):
        dlg = CreateUserDialog(self.db, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self._admin_refresh()

    def _admin_reset_pw(self):
        row = self.admin_table.currentRow()
        if row < 0:
            return
        uid   = int(self.admin_table.item(row, 0).text())
        email = self.admin_table.item(row, 1).text()
        from PyQt6.QtWidgets import QInputDialog
        pw, ok = QInputDialog.getText(self, "Reset Password",
                                      f"New password for {email}:",
                                      QLineEdit.EchoMode.Password)
        if ok and pw:
            if len(pw) < 8:
                QMessageBox.warning(self,"Error","Min 8 characters.")
                return
            self.db.reset_password(uid, pw)
            QMessageBox.information(self, "Done", f"Password reset for {email}.")

    def _admin_toggle_active(self):
        row = self.admin_table.currentRow()
        if row < 0:
            return
        uid    = int(self.admin_table.item(row, 0).text())
        active = self.admin_table.item(row, 4).text() == "Active"
        if uid == self.user["id"]:
            QMessageBox.warning(self, "Error", "Cannot deactivate yourself.")
            return
        self.db.set_user_active(uid, not active)
        self._admin_refresh()

    def _admin_set_subscription(self):
        row = self.admin_table.currentRow()
        if row < 0:
            return
        uid = int(self.admin_table.item(row, 0).text())
        email = self.admin_table.item(row, 1).text()
        role = self.admin_table.item(row, 3).text()
        if role == "admin":
            QMessageBox.information(
                self, "Administrator",
                "Administrator accounts do not expire.")
            return

        from PyQt6.QtWidgets import QInputDialog
        plan, ok = QInputDialog.getItem(
            self, "Subscription", f"Plan for {email}:",
            ["15 days", "Weekly", "Monthly", "No expiry"], 0, False)
        if not ok:
            return
        current_slots = int(self.admin_table.item(row, 7).text() or "1")
        slots, ok = QInputDialog.getInt(
            self, "Max live slots", "Allowed simultaneous slots:",
            current_slots, 1, 5)
        if not ok:
            return
        days = {"15 days": 15, "Weekly": 7, "Monthly": 30}.get(plan)
        expires_at = ""
        if days:
            expires_at = (
                datetime.now() + timedelta(days=days)
            ).strftime("%Y-%m-%d %H:%M:%S")
        self.db.set_user_subscription(uid, plan, expires_at, slots)
        self.db.set_user_active(uid, True)
        self._admin_refresh()

    def _admin_delete(self):
        row = self.admin_table.currentRow()
        if row < 0:
            return
        uid   = int(self.admin_table.item(row, 0).text())
        email = self.admin_table.item(row, 1).text()
        if uid == self.user["id"]:
            QMessageBox.warning(self, "Error", "Cannot delete yourself.")
            return
        if QMessageBox.question(
                self, "Delete User", f"Delete {email}?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        ) == QMessageBox.StandardButton.Yes:
            self.db.delete_user(uid)
            self._admin_refresh()

    # ── Theme / settings ──────────────────────────────────────────────────────

    def _toggle_theme(self):
        current = self.cfg.get("theme", DEFAULT_THEME)
        self._set_theme("light" if current == "dark" else "dark")

    def _sync_dial_buttons(self, running: bool) -> None:
        for btn_start in (getattr(self, "btn_start", None),
                          getattr(self, "btn_start_live", None)):
            if btn_start is not None:
                btn_start.setEnabled(not running)
        for btn_stop in (getattr(self, "btn_stop", None),
                         getattr(self, "btn_stop_live", None)):
            if btn_stop is not None:
                btn_stop.setEnabled(running)
        for btn_stop in (getattr(self, "btn_emergency_stop", None),
                         getattr(self, "btn_emergency_live", None)):
            if btn_stop is not None:
                btn_stop.setEnabled(running)

    def _on_tab_changed(self, index: int) -> None:
        widget = self.tabs.widget(index)
        if widget is getattr(self, "tab_logs", None):
            self._refresh_logs()
        elif widget is getattr(self, "tab_settings", None):
            self._refresh_gv_accounts()

    def _set_theme(self, name: str):
        self.cfg["theme"] = name
        _save_cfg(self.cfg)
        app = QApplication.instance()
        if app is not None:
            apply_theme(app, name)
        self._theme_btn.setText(
            "Light mode" if name == "dark" else "Dark mode")
        if hasattr(self, "log_table"):
            self._apply_log_filter()
        if hasattr(self, "gv_accounts_table"):
            self._refresh_gv_accounts()

    def _save_settings(self):
        self.cfg["n_slots"] = self.settings_slots.value()
        if hasattr(self, "chk_ai_audio"):
            self.cfg["enable_ai_audio"] = self.chk_ai_audio.isChecked()
        if hasattr(self, "audio_device_combo"):
            self.cfg["audio_device"] = self.audio_device_combo.currentData() or ""
        if hasattr(self, "chk_live_debug"):
            self.cfg["live_debug_mode"] = self.chk_live_debug.isChecked()
        _save_cfg(self.cfg)
        if hasattr(self, "_slot_cards"):
            for card in self._slot_cards.values():
                card.set_ai_audio_enabled(bool(self.cfg.get("enable_ai_audio", True)))
        QMessageBox.information(self, "Saved", "Settings saved.")

    # ── Utility ───────────────────────────────────────────────────────────────

    def _log(self, msg: str):
        if hasattr(self, "console"):
            ts = datetime.now().strftime("%H:%M:%S")
            self.console.append(f"[{ts}]  {msg}")

    def _logout(self):
        if self._running:
            if QMessageBox.question(
                    self, "Logout", "Dialer is running. Stop and logout?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            ) != QMessageBox.StandardButton.Yes:
                return
            self._stop_dialing()
        self._watchdog.stop()
        for ctrl in self._controllers:
            if ctrl is not None:
                self._dispose_controller(ctrl)
        self._controllers.clear()
        QApplication.processEvents()
        log_info("User logged out")
        self.close()
        self._app_ref.show_login()

    def set_app_ref(self, app: "DialerApp"):
        self._app_ref = app


# ══════════════════════════════════════════════════════════════════════════════
#  ROOT CONTROLLER
# ══════════════════════════════════════════════════════════════════════════════

class DialerApp:
    def __init__(self):
        self.db  = CRMDatabase()
        self.cfg = _load_cfg()
        self._client_mode = is_client_deployment(self.cfg)
        self._main_win: MainWindow | None = None
        self._stack = QStackedWidget()
        self._stack.setWindowTitle(APP_NAME)
        self._stack.setWindowIcon(_icon())
        self._stack.resize(1000, 680)

        theme = self.cfg.get("theme", DEFAULT_THEME)
        app = QApplication.instance()
        if app is not None:
            apply_theme(app, theme)

        self._route_startup()

        self._stack.show()

    def _route_startup(self) -> None:
        self._bootstrap_admin_from_env()
        if self._client_mode:
            if not self.db.has_any_user():
                self._show_client_not_configured()
            else:
                self._show_login()
            return
        if self.db.needs_admin_setup():
            self._show_admin_setup()
        else:
            self._show_login()

    def _bootstrap_admin_from_env(self) -> None:
        email, password = _admin_env_credentials()
        if not email or not password or not self.db.needs_admin_setup():
            return
        if len(password) < 8:
            log_warning("ADMIN_PASSWORD is set but too short for admin bootstrap")
            return
        try:
            self.db.create_admin(email, "Environment Admin", password)
            log_info(f"Admin account bootstrapped from ADMIN_EMAIL={email}")
        except Exception:
            log_error("Admin env bootstrap failed\n" + traceback.format_exc())

    def _show_admin_setup(self):
        page = AdminSetupPage(self.db)
        page.done.connect(self._show_login)
        self._stack.addWidget(page)
        self._stack.setCurrentWidget(page)

    def _show_client_not_configured(self):
        page = ClientNotConfiguredPage()
        self._stack.addWidget(page)
        self._stack.setCurrentWidget(page)

    def _show_login(self):
        page = LoginPage(self.db, client_mode=self._client_mode)
        page.login_success.connect(self._on_login)
        self._stack.addWidget(page)
        self._stack.setCurrentWidget(page)

    def _on_login(self, user: dict):
        self._stack.hide()
        win = MainWindow(self.db, user, self.cfg)
        win.set_app_ref(self)
        self._main_win = win
        win.show()

    def show_login(self):
        if self._main_win:
            self._main_win = None
        self._stack.resize(1000, 680)
        self._show_login()
        self._stack.show()


# ── Utility ───────────────────────────────────────────────────────────────────
import time as _time
def _now() -> float:
    return _time.time()


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument("--live-test-live-test", action="store_true")
    args, remaining = parser.parse_known_args()
    if args.live_test_live_test:
        from scripts.live_call_smoke import main as live_smoke_main
        sys.argv = [sys.argv[0], "--live-test-live-test", *remaining]
        live_smoke_main()
        raise SystemExit(0)

    _load_dotenv_file(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
    os.makedirs(LOGS_DIR, exist_ok=True)
    cleanup_stale_webengine_processes()

    app = QApplication(sys.argv)
    app_lock = QLockFile(os.path.join(LOGS_DIR, "ftsolutions_autodialer.lock"))
    app_lock.setStaleLockTime(30_000)
    if not app_lock.tryLock(800):
        cleanup_stale_webengine_processes()
        app_lock.removeStaleLockFile()
        if not app_lock.tryLock(800):
            QMessageBox.warning(
                None,
                "FT Solutions Auto Dialer is already open",
                "Close the existing Auto Dialer window before opening another one.\n\n"
                "If no window is visible, open Task Manager and end "
                "QtWebEngineProcess.exe, then try again.",
            )
            sys.exit(1)
    app.setApplicationName("FTSolutions AutoDialer")
    app.setOrganizationName("FT Solutions")
    setup_dialer_logging()

    dialer = DialerApp()
    sys.exit(app.exec())
